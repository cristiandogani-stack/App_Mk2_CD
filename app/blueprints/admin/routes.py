from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, current_app, session
from flask_login import login_required, current_user
from ...extensions import login_manager
from ...extensions import db
from ...models import (
    Module, StructureType, Structure, User, Supplier, WorkCenter, WorkPhase,
    MaterialCost, Product, ProductComponent, ComponentMaster, BOMLine,
    TypeField, ComponentFieldValue, CustomField, CustomValue,
    StockItem, Document, ScanEvent, ProductBuild, ProductBuildItem,
    ProductionBox, InventoryLog, Reservation
)
from types import SimpleNamespace

import os
import shutil  # used for copying images and documents when copying defaults
from werkzeug.utils import secure_filename
from flask import send_file

# Import checklist loader for document flagging
from ...checklist import load_checklist
import tempfile
import zipfile
import io

# -----------------------------------------------------------------------------
# Component master helpers
#
# To support absolute uniqueness of component codes across the application
# structures and product components now reference a ComponentMaster.  The
# helper functions defined below create or look up the master for a given
# Structure, copy attributes and attachments from the node to the master
# record, and assign the master id back to the node.  These helpers are
# called during node creation and editing to ensure data remains
# consistent.

def _copy_attachments_to_master(struct_obj: Structure, master_obj: ComponentMaster):
    """Copy images and document folders from a structure node to its master.

    Images stored under ``static/uploads`` beginning with ``sn_<struct_id>_``
    are copied to files named ``cm_<master_id>_<suffix>``.  Document
    directories in ``static/tmp_structures/<type>/<path>`` are mirrored into
    ``static/tmp_components/<code>`` using the component code as the folder
    name.  Existing files in the destination are preserved.
    """
    try:
        # Copy images
        upload_dir = os.path.join(current_app.static_folder, 'uploads')
        if os.path.isdir(upload_dir):
            prefix_src = f"sn_{struct_obj.id}_"
            for fname in os.listdir(upload_dir):
                if fname.startswith(prefix_src):
                    suffix = fname[len(prefix_src):]
                    dest_name = f"cm_{master_obj.id}_{suffix}"
                    src_path = os.path.join(upload_dir, fname)
                    dest_path = os.path.join(upload_dir, dest_name)
                    try:
                        if not os.path.exists(dest_path):
                            shutil.copyfile(src_path, dest_path)
                    except Exception:
                        pass
        # Copy documents
        # Build source path using the original tmp_structures hierarchy
        def _safe(n: str) -> str:
            return secure_filename(n) or 'unnamed'
        try:
            type_dir = _safe(struct_obj.type.name)
        except Exception:
            type_dir = 'unknown'
        parts: list[str] = []
        def _collect(node_obj):
            if node_obj.parent:
                _collect(node_obj.parent)
            parts.append(_safe(node_obj.name))
        _collect(struct_obj)
        src_base = os.path.join(current_app.static_folder, 'tmp_structures', type_dir, *parts)
        if os.path.isdir(src_base):
            dest_base = os.path.join(current_app.static_folder, 'tmp_components', _safe(master_obj.code))
            try:
                # Create destination directory if it does not exist
                os.makedirs(dest_base, exist_ok=True)
            except Exception:
                pass
            # Copy each subdirectory and file into the destination preserving structure
            for root, dirs, files in os.walk(src_base):
                rel = os.path.relpath(root, src_base)
                dest_root = os.path.join(dest_base, rel) if rel != '.' else dest_base
                try:
                    os.makedirs(dest_root, exist_ok=True)
                except Exception:
                    pass
                for f in files:
                    src_file = os.path.join(root, f)
                    dest_file = os.path.join(dest_root, f)
                    try:
                        if not os.path.exists(dest_file):
                            shutil.copyfile(src_file, dest_file)
                    except Exception:
                        pass
    except Exception:
        # Fail silently on any error
        pass

def ensure_component_master_for_structure(struct_obj: Structure) -> ComponentMaster:
    """Ensure that a given structure has an associated ComponentMaster.

    This function looks up a master record by the structure name (treated as
    the component code).  If none exists, a new ComponentMaster is created
    and populated with attributes copied from the structure.  The structure's
    ``component_id`` is updated to reference the master.  Attachments (images
    and documents) are copied from the structure directories into the master
    directories.  The function returns the master record.
    """
    if not struct_obj:
        return None
    # Derive component code from the structure name
    code = (struct_obj.name or '').strip()
    if not code:
        return None
    # Search for existing master
    master = ComponentMaster.query.filter_by(code=code).first()
    created = False
    if not master:
        # Create new master and copy attributes from the structure
        master = ComponentMaster(code=code)
        # Copy numeric and textual fields from the structure as the initial
        # canonical values.  Use getattr with default None to avoid errors.
        master.weight = struct_obj.weight
        master.description = struct_obj.description
        master.notes = struct_obj.notes
        master.processing_type = struct_obj.processing_type
        master.work_phase_id = struct_obj.work_phase_id
        master.supplier_id = struct_obj.supplier_id
        master.work_center_id = struct_obj.work_center_id
        master.processing_cost = struct_obj.processing_cost
        master.standard_time = struct_obj.standard_time
        master.lead_time_theoretical = struct_obj.lead_time_theoretical
        master.lead_time_real = struct_obj.lead_time_real
        master.price_per_unit = struct_obj.price_per_unit
        master.minimum_order_qty = struct_obj.minimum_order_qty
        # cycles_json and type_of_processing are left None unless set via custom logic
        db.session.add(master)
        db.session.commit()
        created = True
    # Assign the structure's component_id if not already assigned
    if struct_obj.component_id != master.id:
        struct_obj.component_id = master.id
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
    # Copy attachments if the master was newly created or if the structure's
    # attachments are newer.  We copy unconditionally to simplify logic.
    _copy_attachments_to_master(struct_obj, master)
    return master

admin_bp = Blueprint('admin', __name__, template_folder='../../templates')

# -----------------------------------------------------------------------------
# Runtime schema upgrade helpers
#
# The following helper function ensures that newly introduced columns for
# inventory management exist in the database.  Because this simplified
# application does not use a formal migration tool like Alembic, missing
# columns are added on the fly using ALTER TABLE statements.  The checks are
# idempotent and safe to run on every request; when the columns already
# exist, the ALTER statements are skipped.  Using PRAGMA table_info,
# we inspect the current schema of each table and conditionally add the
# stock_threshold and replenishment_qty columns for the relevant tables.

def _ensure_inventory_columns() -> None:
    """Add inventory-related columns to tables if they do not exist."""
    try:
        # Import here to avoid circular import issues when Flask initializes
        from sqlalchemy import text
        conn = db.engine.connect()
        # structure_types table: add default_stock_threshold and default_replenishment_qty
        res = conn.execute(text('PRAGMA table_info(structure_types)')).fetchall()
        col_names = [row[1] for row in res]
        if 'default_stock_threshold' not in col_names:
            conn.execute(text('ALTER TABLE structure_types ADD COLUMN default_stock_threshold FLOAT'))
        if 'default_replenishment_qty' not in col_names:
            conn.execute(text('ALTER TABLE structure_types ADD COLUMN default_replenishment_qty FLOAT'))
        # structures table: add stock_threshold, replenishment_qty, revision and
        # compatible_revisions.  The compatible_revisions column stores a
        # comma‑separated list of revision labels selected as compatible when
        # revising a structure.  It is added dynamically to ensure older
        # databases remain backwards compatible.
        res2 = conn.execute(text('PRAGMA table_info(structures)')).fetchall()
        col2 = [row[1] for row in res2]
        # Add inventory stock and replenishment fields when missing
        if 'stock_threshold' not in col2:
            conn.execute(text('ALTER TABLE structures ADD COLUMN stock_threshold FLOAT'))
        if 'replenishment_qty' not in col2:
            conn.execute(text('ALTER TABLE structures ADD COLUMN replenishment_qty FLOAT'))
        # Add the revision column when absent.  The revision column stores
        # the numeric revision index for each structure.  It is initialised
        # to 0 for existing rows so that the UI displays an empty
        # string until the first revision is made.  This dynamic
        # migration enables backwards compatibility with databases
        # created before the revision feature was introduced.
        if 'revision' not in col2:
            conn.execute(text('ALTER TABLE structures ADD COLUMN revision INTEGER DEFAULT 0'))
        # Add the compatible_revisions column when absent.  This column
        # stores a comma‑separated list of revision labels (e.g. "Rev.A,Rev.B")
        # indicating which prior versions remain compatible after a
        # revision.  When missing we add it as a TEXT column.
        if 'compatible_revisions' not in col2:
            conn.execute(text('ALTER TABLE structures ADD COLUMN compatible_revisions TEXT'))
        # product_components table: add stock_threshold and replenishment_qty
        res3 = conn.execute(text('PRAGMA table_info(product_components)')).fetchall()
        col3 = [row[1] for row in res3]
        if 'stock_threshold' not in col3:
            conn.execute(text('ALTER TABLE product_components ADD COLUMN stock_threshold FLOAT'))
        if 'replenishment_qty' not in col3:
            conn.execute(text('ALTER TABLE product_components ADD COLUMN replenishment_qty FLOAT'))
        # component_masters table: add stock_threshold and replenishment_qty
        res4 = conn.execute(text('PRAGMA table_info(component_masters)')).fetchall()
        col4 = [row[1] for row in res4]
        if 'stock_threshold' not in col4:
            conn.execute(text('ALTER TABLE component_masters ADD COLUMN stock_threshold FLOAT'))
        if 'replenishment_qty' not in col4:
            conn.execute(text('ALTER TABLE component_masters ADD COLUMN replenishment_qty FLOAT'))
        conn.close()
    except Exception:
        # Silently ignore any errors during schema upgrades.  In production,
        # logging would be appropriate to detect issues.
        pass

def admin_required():
    """Ensure the current user is logged in and has admin privileges.

    This helper raises a 403 error when a non‑admin user attempts to access an
    administrative route.  It does not handle redirecting anonymous users
    because Flask‑Login's `login_required` decorator takes care of that.
    """
    if not current_user.is_authenticated or not getattr(current_user, 'is_admin', False):
        abort(403)

@admin_bp.before_request
def check_admin():
    """Protect all admin routes so that only authenticated admin users may access them.

    Flask‑Login’s `login_required` decorator on individual view functions will
    handle redirecting unauthenticated visitors to the login page.  For
    authenticated users we further verify they have the admin role.  If not,
    we abort with a 403 Forbidden response.  The `request.endpoint` check is
    unnecessary because this hook only applies to this blueprint.
    """
    if not current_user.is_authenticated:
        # Delegate to Flask‑Login for redirect behaviour
        return login_manager.unauthorized()
    if not current_user.is_admin:
        abort(403)

@admin_bp.route('/')
@login_required
def index():
    admin_required()
    return render_template('admin/index.html')

@admin_bp.route('/modules', methods=['GET', 'POST'])
@login_required
def modules():
    admin_required()
    if request.method == 'POST':
        module_id = request.form.get('module_id')
        action = request.form.get('action')
        m = Module.query.get(int(module_id))
        if not m:
            flash('Modulo non trovato', 'warning')
            return redirect(url_for('admin.modules'))
        if action == 'toggle':
            m.enabled = not m.enabled
            db.session.commit()
            flash(f'Modulo "{m.name}": {"abilitato" if m.enabled else "disabilitato"}', 'success')
        return redirect(url_for('admin.modules'))

    modules = Module.query.order_by(Module.name.asc()).all()
    return render_template('admin/modules.html', modules=modules)


# -----------------------------------------------------------------------------
# Database administration endpoints
# These endpoints allow administrators to reset, save (export) and update
# (import) the application database along with its attachments.  The reset
# operation wipes all structures, products, product components and component
# masters as well as uploaded images and temporary documents.  The save
# operation packages the current database file (instance/app.db) together
# with the contents of ``static/uploads``, ``static/tmp_structures`` and
# ``static/tmp_components`` into a ZIP archive for download.  The update
# operation accepts a ZIP file previously created by the save endpoint and
# restores the database and attachments from it.

@admin_bp.route('/db/reset', methods=['POST'])
@login_required
def reset_db():
    """Erase all structures, products and attachments from the database.

    This action removes all entries from ProductComponent, BOMLine, Product,
    Structure and ComponentMaster tables.  It also deletes uploaded images
    (except for static assets like logos) and temporary document directories.
    Only accessible by administrators.  A confirmation should be presented
    to the user on the client side before invoking this endpoint.
    """
    admin_required()
    # Delete database records
    try:
        # Remove dependent records in an order that avoids foreign key conflicts.
        # Use session-level deletes to ensure cascades are honored.
        # Define the models to purge.  This list includes all tables that
        # store inventory, registry and build data.  Exclude User and Module
        # models so that admin accounts and module settings persist.
        for model in [
            # Inventory and build records (children first to satisfy foreign keys)
            ProductBuildItem,  # depends on ProductBuild and Product
            StockItem,         # depends on Product, Reservation, ProductionBox
            Document,          # depends on StockItem or ProductionBox
            ScanEvent,         # loosely coupled; remove early
            InventoryLog,      # references structures
            ProductBuild,      # depends on Product
            ProductionBox,     # parent of StockItem
            Reservation,       # parent of StockItem
            # Custom and dictionary tables and BOM definitions
            ProductComponent, BOMLine,
            # Core entities
            Product, Structure, StructureType,
            Supplier, WorkCenter, WorkPhase, MaterialCost, ComponentMaster,
            # Custom field definitions and values last
            ComponentFieldValue, TypeField, CustomValue, CustomField
        ]:
            try:
                db.session.query(model).delete(synchronize_session=False)
            except Exception:
                # Fallback: iterate individual objects for deletion
                for obj in model.query.all():
                    try:
                        db.session.delete(obj)
                    except Exception:
                        pass
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Errore durante la cancellazione dei dati.', 'danger')
        return redirect(url_for('admin.modules'))
    # Remove uploaded images except for built-in assets (logo, favicon)
    try:
        upload_dir = os.path.join(current_app.static_folder, 'uploads')
        if os.path.isdir(upload_dir):
            for fname in os.listdir(upload_dir):
                # Skip known assets (if any) by extension; remove all others
                if fname.lower().startswith('logo') or fname.lower().startswith('favicon'):
                    continue
                fpath = os.path.join(upload_dir, fname)
                try:
                    os.remove(fpath)
                except Exception:
                    pass
    except Exception:
        # ignore errors cleaning uploads
        pass
    # Remove temporary document directories
    for sub in ['tmp_structures', 'tmp_components']:
        try:
            base = os.path.join(current_app.static_folder, sub)
            if os.path.isdir(base):
                shutil.rmtree(base, ignore_errors=True)
        except Exception:
            pass
    # Remove all user uploaded documents under static/documents
    try:
        docs_dir = os.path.join(current_app.static_folder, 'documents')
        if os.path.isdir(docs_dir):
            for entry in os.listdir(docs_dir):
                path_entry = os.path.join(docs_dir, entry)
                try:
                    if os.path.isdir(path_entry):
                        shutil.rmtree(path_entry, ignore_errors=True)
                    else:
                        os.remove(path_entry)
                except Exception:
                    pass
    except Exception:
        pass
    # Remove production archives and completed assemblies (under Produzione)
    try:
        prod_base = os.path.join(current_app.root_path, 'Produzione')
        for sub_name in ['archivio', 'Assiemi_completati']:
            d = os.path.join(prod_base, sub_name)
            if os.path.isdir(d):
                shutil.rmtree(d, ignore_errors=True)
                try:
                    os.makedirs(d, exist_ok=True)
                except Exception:
                    pass
    except Exception:
        pass
    # Remove the checklist file so that all previous document flags are cleared
    try:
        from ...checklist import _get_checklist_path  # Local import to avoid circular import
        cl_path = _get_checklist_path()
        if cl_path and os.path.exists(cl_path):
            os.remove(cl_path)
    except Exception:
        pass
    flash('Database resettato: tutti i dati e file sono stati rimossi.', 'success')
    return redirect(url_for('admin.modules'))


@admin_bp.route('/db/save', methods=['POST'])
@login_required
def save_db():
    """Package the current database and attachments into a ZIP archive.

    This endpoint copies the SQLite database file ``instance/app.db`` and the
    directories ``static/uploads``, ``static/tmp_structures`` and
    ``static/tmp_components`` into a temporary location, zips them and
    returns the resulting archive for download.  The archive can later be
    imported via the update_db endpoint.  Only accessible by administrators.
    """
    admin_required()
    try:
        # Create a temporary directory to assemble the backup
        with tempfile.TemporaryDirectory() as tmpdir:
            # Copy the database file
            src_db = os.path.join(current_app.instance_path, 'app.db')
            dest_db = os.path.join(tmpdir, 'app.db')
            try:
                shutil.copyfile(src_db, dest_db)
            except Exception:
                # If copy fails, still proceed but warn user
                pass
            # Copy attachments directories
            static_base = current_app.static_folder
            for sub in ['uploads', 'tmp_structures', 'tmp_components']:
                src = os.path.join(static_base, sub)
                if os.path.isdir(src):
                    dest = os.path.join(tmpdir, sub)
                    try:
                        shutil.copytree(src, dest, dirs_exist_ok=True)
                    except Exception:
                        pass
            # Create an in-memory zip file
            mem_file = io.BytesIO()
            with zipfile.ZipFile(mem_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(tmpdir):
                    for filename in files:
                        abs_path = os.path.join(root, filename)
                        rel_path = os.path.relpath(abs_path, tmpdir)
                        try:
                            zf.write(abs_path, rel_path)
                        except Exception:
                            pass
            mem_file.seek(0)
            return send_file(mem_file, as_attachment=True, download_name='db_backup.zip', mimetype='application/zip')
    except Exception:
        flash('Errore durante il salvataggio del database.', 'danger')
        return redirect(url_for('admin.modules'))


@admin_bp.route('/db/update', methods=['GET', 'POST'])
@login_required
def update_db():
    """Import database and attachments from a previously exported ZIP archive.

    On GET this view displays a simple file upload form.  On POST it
    extracts the uploaded archive, replaces the current database file and
    attachments directories, and reloads data.  Only accessible by
    administrators.
    """
    admin_required()
    if request.method == 'POST':
        file = request.files.get('db_file')
        if not file or not file.filename:
            flash('Nessun file selezionato.', 'warning')
            return redirect(url_for('admin.update_db'))
        # Save the uploaded file to a temporary file
        try:
            tmp_upload = tempfile.NamedTemporaryFile(delete=False)
            file.save(tmp_upload.name)
        except Exception:
            flash('Errore durante il caricamento del file.', 'danger')
            return redirect(url_for('admin.update_db'))
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                # Extract the archive
                with zipfile.ZipFile(tmp_upload.name, 'r') as zf:
                    zf.extractall(tmpdir)
                # Replace database
                src_db = os.path.join(tmpdir, 'app.db')
                if os.path.isfile(src_db):
                    dest_db = os.path.join(current_app.instance_path, 'app.db')
                    try:
                        shutil.copyfile(src_db, dest_db)
                    except Exception:
                        flash('Impossibile sostituire il file del database.', 'danger')
                        return redirect(url_for('admin.update_db'))
                # Replace attachments directories
                static_base = current_app.static_folder
                for sub in ['uploads', 'tmp_structures', 'tmp_components']:
                    src_dir = os.path.join(tmpdir, sub)
                    if os.path.isdir(src_dir):
                        dest_dir = os.path.join(static_base, sub)
                        try:
                            # Remove existing directory if any
                            if os.path.isdir(dest_dir):
                                shutil.rmtree(dest_dir, ignore_errors=True)
                            shutil.copytree(src_dir, dest_dir)
                        except Exception:
                            pass
        finally:
            # Clean up the temporary uploaded file
            try:
                os.unlink(tmp_upload.name)
            except Exception:
                pass
        flash('Database aggiornato con successo.', 'success')
        return redirect(url_for('admin.modules'))
    # GET: display upload form
    return render_template('admin/update_db.html')

@admin_bp.route('/structures', methods=['GET', 'POST'])
@login_required
def structures():
    admin_required()
    # Ensure inventory columns exist before querying.  Without this check, newly
    # added fields such as default_stock_threshold on the structure_types table
    # may cause an OperationalError on SQLite when the columns are missing.
    _ensure_inventory_columns()
    if request.method == 'POST':
        form_type = request.form.get('form_type')
        action = request.form.get('action') or 'create'
        # Handle creation or definition for structure types
        if form_type == 'type':
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            # Determine typology from radio input.  When the typology field is omitted
            # (e.g. because no radio buttons are displayed), all flags default to False.
            typology = request.form.get('typology') or None
            is_ass = typology == 'assembly'
            is_part = typology == 'part'
            is_comm = typology == 'commercial'
            if not name:
                flash('Il nome è obbligatorio.', 'danger')
                return redirect(url_for('admin.structures'))
            # Persist the entered values in the session so that the form remains populated after redirects
            session['pending_type_form'] = {
                'name': name,
                'description': description
            }
            # If the form still included a typology selection, remember it for the next request
            if typology:
                session['pending_type_form']['typology'] = typology
            # If a type with the same name already exists, reuse it when defining
            existing_type = StructureType.query.filter_by(name=name).first()
            # If the action is "define", reuse an existing type (if any) instead of creating a duplicate
            if action == 'define':
                # When a type with this name exists, simply redirect to its definition page
                if existing_type:
                    return redirect(url_for('admin.define_structure_type_defaults', type_id=existing_type.id))
                # Otherwise, create a new type and allow defining defaults
                st = StructureType(
                    name=name,
                    description=description,
                    is_assembly=is_ass,
                    is_part=is_part,
                    is_commercial=is_comm
                )
                db.session.add(st)
                db.session.commit()
                return redirect(url_for('admin.define_structure_type_defaults', type_id=st.id))
            # For plain creation (not define), ensure the type does not already exist
            if existing_type:
                flash('Esiste già un tipo con questo nome.', 'warning')
                return redirect(url_for('admin.structures'))
            # Create the type and clear pending form on successful creation
            st = StructureType(
                name=name,
                description=description,
                is_assembly=is_ass,
                is_part=is_part,
                is_commercial=is_comm
            )
            db.session.add(st)
            db.session.commit()
            # When a new structure type is created, also create a corresponding
            # product if one does not already exist.  This aligns the domain
            # model where each typology represents a distinct product.  The
            # product inherits the name and description from the type and
            # initialises other fields (revision, revisionabile) with default
            # values.  Avoid duplicating products when a type with the same
            # name is created again.
            try:
                existing_product = Product.query.filter_by(name=name).first()
            except Exception:
                existing_product = None
            if not existing_product:
                try:
                    new_prod = Product(
                        name=name,
                        description=description,
                        revision=1,
                        revisionabile=False
                    )
                    db.session.add(new_prod)
                    db.session.commit()
                except Exception:
                    # On failure (e.g. missing table) roll back the transaction but keep the type
                    db.session.rollback()
            # Clear any pending form data and notify the admin
            session.pop('pending_type_form', None)
            flash('Tipo creato.', 'success')
        # Handle creation or definition for structure nodes
        elif form_type == 'node':
            type_id = request.form.get('type_id')
            parent_id = request.form.get('parent_id') or None
            name = request.form.get('node_name', '').strip()
            node_typology = request.form.get('node_typology') or ''
            flag_ass = node_typology == 'assembly'
            flag_part = node_typology == 'part'
            flag_comm = node_typology == 'commercial'
            if not type_id or not name:
                flash('Compila tutti i campi obbligatori.', 'danger')
                return redirect(url_for('admin.structures'))
            if not node_typology:
                flash('Seleziona una tipologia per il nodo (Assieme, Parte o Parte a commercio).', 'danger')
                return redirect(url_for('admin.structures'))
            # Persist the entered node form values so that the form remains populated after redirects
            session['pending_node_form'] = {
                'type_id': type_id,
                'parent_id': parent_id,
                'node_name': name,
                'node_typology': node_typology
            }
            # Determine which draft flag to use based on typology
            draft_field = None
            if flag_ass:
                draft_field = 'table_assieme'
            elif flag_part:
                draft_field = 'table_parte'
            else:
                draft_field = 'table_commerciale'
            # Helper to finalise an existing draft node
            def finalize_draft(draft):
                if flag_ass:
                    draft.table_assieme = False
                elif flag_part:
                    draft.table_parte = False
                else:
                    draft.table_commerciale = False
                db.session.commit()
                return draft
            # When creating a node (action = create) check for an existing draft
            if action != 'define':
                # When creating a new node (not define) allow multiple structures with the
                # same name and typology across different types.  To ensure that
                # nodes imported or created with the same code reuse existing
                # details (description, notes, cycles, etc.), we create a new
                # structure for the requested type/parent but copy attributes from
                # the first non‑draft structure with the same name and typology.
                # Draft nodes are handled separately below.
                # See if there is a draft node with same name, type and parent and typology
                existing_draft = (Structure.query
                                  .filter_by(name=name, type_id=int(type_id), parent_id=int(parent_id) if parent_id else None,
                                             flag_assembly=flag_ass, flag_part=flag_part, flag_commercial=flag_comm)
                                  .filter(getattr(Structure, draft_field) == True)
                                  .first())
                if existing_draft:
                    # Finalize the draft instead of creating a new node
                    s = finalize_draft(existing_draft)
                else:
                    s = Structure(
                        name=name,
                        type_id=int(type_id),
                        parent_id=int(parent_id) if parent_id else None,
                        flag_assembly=flag_ass,
                        flag_part=flag_part,
                        flag_commercial=flag_comm
                    )
                    db.session.add(s)
                    db.session.commit()
                    # After creating the new structure, copy attributes from an
                    # existing non‑draft structure with the same name and typology
                    # (if any).  This ensures that details (weight, processing,
                    # description, notes, cycles, etc.) are reused when the
                    # same code is imported into multiple types.
                    existing_global = (Structure.query
                                       .filter_by(name=name,
                                                  flag_assembly=flag_ass,
                                                  flag_part=flag_part,
                                                  flag_commercial=flag_comm)
                                       .filter_by(table_assieme=False,
                                                  table_parte=False,
                                                  table_commerciale=False)
                                       .filter(Structure.id != s.id)
                                       .first())
                    if existing_global:
                        # Copy attributes that are defined on the canonical structure to the new
                        # structure only if they are not already set.  We avoid copying
                        # identifiers like type_id and parent_id.  Attributes include
                        # weight, processing type, work phase, supplier, work centre,
                        # times, descriptions, notes, price, MOQ and processing cost.
                        def copy_attr(attr):
                            return getattr(existing_global, attr)
                        attrs = [
                            'weight', 'processing_type', 'work_phase_id', 'supplier_id', 'work_center_id',
                            'standard_time', 'lead_time_theoretical', 'lead_time_real', 'description', 'notes',
                            'price_per_unit', 'minimum_order_qty', 'processing_cost'
                        ]
                        updated = False
                        for attr in attrs:
                            if getattr(s, attr) in (None, '', 0) and getattr(existing_global, attr) not in (None, ''):
                                setattr(s, attr, copy_attr(attr))
                                updated = True
                        if updated:
                            try:
                                db.session.commit()
                            except Exception:
                                db.session.rollback()
                        # If a canonical node with the same name exists, replicate its
                        # attachments (image and document folders) to the newly created
                        # node.  This ensures that when multiple nodes share the same
                        # code, the new node inherits all files associated with the
                        # existing node.
                        try:
                            # Replicate image from existing_global to new node ``s``
                            upload_dir = os.path.join(current_app.static_folder, 'uploads')
                            if existing_global and os.path.isdir(upload_dir):
                                prefix_src = f"sn_{existing_global.id}_"
                                src_file = None
                                for _f in os.listdir(upload_dir):
                                    if _f.startswith(prefix_src):
                                        src_file = _f
                                        break
                                if src_file:
                                    suffix = src_file[len(prefix_src):]
                                    src_path = os.path.join(upload_dir, src_file)
                                    dest_name = f"sn_{s.id}_{suffix}"
                                    dest_path = os.path.join(upload_dir, dest_name)
                                    try:
                                        shutil.copyfile(src_path, dest_path)
                                    except Exception:
                                        pass
                            # Replicate documents directory from existing_global to s
                            if existing_global:
                                def _safe_name(n: str) -> str:
                                    return secure_filename(n) or 'unnamed'
                                # Build source path parts
                                type_dir_src = _safe_name(existing_global.type.name)
                                parts_src: list[str] = []
                                def _collect_src(node_obj):
                                    if node_obj.parent:
                                        _collect_src(node_obj.parent)
                                    parts_src.append(_safe_name(node_obj.name))
                                _collect_src(existing_global)
                                base_src = os.path.join(current_app.static_folder, 'tmp_structures', type_dir_src, *parts_src)
                                if os.path.isdir(base_src):
                                    # Build destination path parts for new node
                                    type_dir_dst = _safe_name(s.type.name)
                                    parts_dst: list[str] = []
                                    def _collect_dst(node_obj):
                                        if node_obj.parent:
                                            _collect_dst(node_obj.parent)
                                        parts_dst.append(_safe_name(node_obj.name))
                                    _collect_dst(s)
                                    base_dst = os.path.join(current_app.static_folder, 'tmp_structures', type_dir_dst, *parts_dst)
                                    try:
                                        if os.path.isdir(base_dst):
                                            shutil.rmtree(base_dst)
                                        shutil.copytree(base_src, base_dst)
                                    except Exception:
                                        pass
                        except Exception:
                            # Silently ignore replication errors
                            pass
                # ------------------------------------------------------------------
                # Populate newly created or finalised nodes with the default values
                # defined on the associated structure type.  When a new node is
                # created from scratch (i.e. not a draft) it will not inherit
                # attributes from its type unless we do so explicitly here.  The
                # admin may have set default weight, processing parameters,
                # supplier/work centre information, lead times, descriptions and
                # notes (including JSON with additional cycles) on the type via
                # the "Definisci" form.  Copy those values to the node only if
                # they have not already been specified on the node itself.  This
                # ensures that values set during node definition (e.g. via the
                # dedicated defaults page) are not overwritten.
                try:
                    stype_obj = StructureType.query.get(s.type_id)
                except Exception:
                    stype_obj = None
                if stype_obj:
                    # Weight
                    if s.weight is None and stype_obj.default_weight is not None:
                        s.weight = stype_obj.default_weight
                    # Processing type
                    if not s.processing_type and stype_obj.default_processing_type:
                        s.processing_type = stype_obj.default_processing_type
                    # Work phase
                    if s.work_phase_id is None and stype_obj.default_work_phase_id:
                        s.work_phase_id = stype_obj.default_work_phase_id
                    # Supplier
                    if s.supplier_id is None and stype_obj.default_supplier_id:
                        s.supplier_id = stype_obj.default_supplier_id
                    # Work centre
                    if s.work_center_id is None and stype_obj.default_work_center_id:
                        s.work_center_id = stype_obj.default_work_center_id
                    # Standard time (minutes)
                    if s.standard_time is None and stype_obj.default_standard_time is not None:
                        s.standard_time = stype_obj.default_standard_time
                    # Lead times (minutes)
                    if s.lead_time_theoretical is None and stype_obj.default_lead_time_theoretical is not None:
                        s.lead_time_theoretical = stype_obj.default_lead_time_theoretical
                    if s.lead_time_real is None and stype_obj.default_lead_time_real is not None:
                        s.lead_time_real = stype_obj.default_lead_time_real
                    # Description and notes
                    if not s.description and stype_obj.default_description:
                        s.description = stype_obj.default_description
                    if not s.notes and stype_obj.default_notes:
                        s.notes = stype_obj.default_notes
                    # Price per unit and minimum order quantity
                    if s.price_per_unit is None and stype_obj.default_price_per_unit is not None:
                        s.price_per_unit = stype_obj.default_price_per_unit
                    if s.minimum_order_qty is None and stype_obj.default_minimum_order_qty is not None:
                        s.minimum_order_qty = stype_obj.default_minimum_order_qty
                    # Derive processing cost.  Prefer the type default cost when
                    # provided; otherwise compute cost for internal processing
                    # based on standard time and work centre hourly cost.  Reset
                    # processing_cost unconditionally so that we always recalc
                    # from the copied defaults when relevant.
                    s.processing_cost = None
                    try:
                        # If type has an explicit default processing cost, use it
                        if stype_obj.default_processing_cost is not None:
                            s.processing_cost = stype_obj.default_processing_cost
                        else:
                            if s.processing_type == 'internal' and s.standard_time is not None and s.work_center_id:
                                wc_ref = WorkCenter.query.get(s.work_center_id)
                                if wc_ref and wc_ref.hourly_cost is not None:
                                    s.processing_cost = (s.standard_time / 60.0) * float(wc_ref.hourly_cost)
                            elif s.processing_type == 'external' and stype_obj.default_processing_cost is not None:
                                s.processing_cost = stype_obj.default_processing_cost
                    except Exception:
                        # If anything goes wrong computing cost leave it as None
                        s.processing_cost = s.processing_cost
                    # Persist updates to the database
                    try:
                        db.session.commit()
                    except Exception:
                        db.session.rollback()
                # ------------------------------------------------------------------
                # Ensure a ComponentMaster exists for this node and assign it.  This
                # helper will create a master if missing, copy the node's
                # attributes and attachments, and update s.component_id.  It
                # returns the master object for further use.
                try:
                    master_obj = ensure_component_master_for_structure(s)
                except Exception:
                    master_obj = None
                # ------------------------------------------------------------------
                # After finalizing or creating, propagate to existing products
                try:
                    new_node_id = s.id
                    parent_components: list[ProductComponent] = []
                    if parent_id:
                        # When the new node has a parent, propagate the node to all products
                        # that already include the parent structure.  Each ProductComponent
                        # represents a link between a product and its structure node.  By
                        # iterating through these we generate a new ProductComponent for
                        # the child node for each product that contains the parent.
                        parent_id_int = int(parent_id)
                        parent_components = ProductComponent.query.filter_by(structure_id=parent_id_int).all()
                    else:
                        # When the node has no parent (i.e. it is a root node of the type)
                        # we look up all ProductComponent rows referencing the same
                        # structure type.  These represent all existing components of this
                        # type across different products and allow us to propagate the new
                        # node to those products.  If none exist, this indicates that
                        # the structure type has not yet been associated with any product.
                        try:
                            type_int = int(type_id)
                        except Exception:
                            type_int = s.type_id
                        parent_components = (
                            ProductComponent.query
                            .join(Structure)
                            .filter(Structure.type_id == type_int)
                            .all()
                        )
                    # If no parent components are found and the node is a root (no parent_id),
                    # automatically bind the node to the product corresponding to the
                    # structure type.  This ensures that newly created types without any
                    # existing components still have their root nodes visible in the
                    # inventory.  We locate (or create) a product whose name matches
                    # the type's name and attach the new node to it.
                    if not parent_components and not parent_id:
                        # Fetch the structure type via type_id; fall back to the node's type
                        type_obj = StructureType.query.get(type_int)
                        prod = None
                        if type_obj:
                            # Look up a product with the same name as the type
                            try:
                                prod = Product.query.filter_by(name=type_obj.name).first()
                            except Exception:
                                prod = None
                            # If no such product exists, create it with a default revision
                            if not prod:
                                prod = Product(
                                    name=type_obj.name,
                                    description=type_obj.description or '',
                                    revision=1,
                                    revisionabile=False
                                )
                                db.session.add(prod)
                                db.session.flush()
                            # Create a dummy ProductComponent to drive the loop below
                            fake_pc = ProductComponent(product_id=prod.id, structure_id=0)
                            parent_components = [fake_pc]
                    # For each parent component (real or synthetic), create a new
                    # ProductComponent linking the product to this new structure.  This
                    # logic mirrors the existing behaviour for child nodes and also
                    # handles the synthetic entry created above when there are no
                    # existing components for the type.
                    for pc in parent_components:
                        # When a synthetic ProductComponent is used (structure_id == 0),
                        # its structure_id does not correspond to a real structure.  The
                        # goal is only to propagate the new node to the associated product.
                        target_product_id = pc.product_id
                        # Skip synthetic structure_id checks; the subsequent lookup uses
                        # only product_id and new_node_id.
                        existing_pc = ProductComponent.query.filter_by(
                            product_id=target_product_id,
                            structure_id=new_node_id
                        ).first()
                        if not existing_pc:
                            new_pc = ProductComponent(
                                product_id=target_product_id,
                                structure_id=new_node_id,
                                quantity=1
                            )
                            # Always reference the master component if available
                            if s.component_id:
                                new_pc.component_id = s.component_id
                            # Copy default attributes from the structure node to the new component
                            # These become per-product overrides and will generally be ignored
                            # in favour of the master values.
                            if s.weight is not None:
                                new_pc.weight = s.weight
                            if s.processing_type:
                                new_pc.processing_type = s.processing_type
                            if s.work_phase_id:
                                new_pc.work_phase_id = s.work_phase_id
                            if s.supplier_id:
                                new_pc.supplier_id = s.supplier_id
                            if s.work_center_id:
                                new_pc.work_center_id = s.work_center_id
                            if s.standard_time is not None:
                                new_pc.standard_time = s.standard_time
                            if s.lead_time_theoretical is not None:
                                new_pc.lead_time_theoretical = s.lead_time_theoretical
                            if s.lead_time_real is not None:
                                new_pc.lead_time_real = s.lead_time_real
                            if s.processing_cost is not None:
                                new_pc.processing_cost = s.processing_cost
                            if s.description:
                                new_pc.description = s.description
                            if s.notes:
                                new_pc.notes = s.notes
                            if s.price_per_unit is not None:
                                new_pc.price_per_unit = s.price_per_unit
                            if s.minimum_order_qty is not None:
                                new_pc.minimum_order_qty = s.minimum_order_qty
                            db.session.add(new_pc)
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                flash('Nodo creato.', 'success')
                # Clear the pending node form after a successful creation
                session.pop('pending_node_form', None)
                return redirect(url_for('admin.structures'))
            else:
                # Define action: create a draft node, mark as temporary and redirect to defaults
                s = Structure(
                    name=name,
                    type_id=int(type_id),
                    parent_id=int(parent_id) if parent_id else None,
                    flag_assembly=flag_ass,
                    flag_part=flag_part,
                    flag_commercial=flag_comm
                )
                # Mark as draft using the appropriate table flag
                if flag_ass:
                    s.table_assieme = True
                elif flag_part:
                    s.table_parte = True
                else:
                    s.table_commerciale = True
                db.session.add(s)
                db.session.commit()
                return redirect(url_for('admin.define_structure_node_defaults', node_id=s.id))

    types = StructureType.query.order_by(StructureType.name.asc()).all()
    nodes_by_type = {t.id: t.nodes.order_by(Structure.name.asc()).all() for t in types}
    # Fetch dictionary lists used for the definisci fields in the form
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()
    # Include any pending form values in the context so that the forms remain pre‑populated
    pending_type_form = session.get('pending_type_form')
    pending_node_form = session.get('pending_node_form')
    return render_template('admin/structures.html',
                           types=types,
                           nodes_by_type=nodes_by_type,
                           suppliers=suppliers,
                           work_centers=work_centers,
                           work_phases=work_phases,
                           pending_type_form=pending_type_form,
                           pending_node_form=pending_node_form)

# -----------------------------------------------------------------------------
# Default definition pages for structure types and nodes
#
# These views provide dedicated pages to define default attributes for a
# StructureType or Structure.  They are invoked when the administrator
# selects "Definisci" during creation or editing.  The UI closely mirrors
# the product component editing pages and hides the typology radio buttons.

@admin_bp.route('/structures/type/<int:type_id>/defaults', methods=['GET', 'POST'])
@login_required
def define_structure_type_defaults(type_id: int):
    """Display and update default attributes for a structure type.

    When a type is created with action 'define' or a user clicks 'Definisci' from
    the edit page, this view presents a form similar to the product component
    pages.  It allows setting weight, processing parameters, purchase details etc.
    The available fields depend on whether the type is an assembly, part or commercial.
    """
    admin_required()
    # Ensure inventory columns exist prior to performing queries or updates.  This
    # prevents OperationalError exceptions when the database schema lacks
    # the newly introduced default_stock_threshold and default_replenishment_qty columns.
    _ensure_inventory_columns()
    st = StructureType.query.get_or_404(type_id)
    # Determine category based on flags
    if st.is_assembly:
        category = 'assembly'
    elif st.is_part:
        category = 'part'
    else:
        category = 'commercial'

    # Helper function to validate allowed image file extensions.  This mirrors
    # the logic used for product component images.  Administrators may
    # upload a representative image when defining defaults.  Accepted
    # extensions include common raster formats.
    def _allowed_image(filename: str) -> bool:
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in {
            'png', 'jpg', 'jpeg', 'gif', 'bmp'
        }
    if request.method == 'POST':
        """Handle file uploads and default attribute updates for structure types.

        Document uploads are processed on every POST to ensure files are saved
        even when the administrator does not provide an explicit action flag.
        When the submitted form contains only file inputs (e.g. for assemblies)
        we skip updating any default attributes to avoid overwriting existing
        values.  Otherwise attributes and optional cycles are processed.
        """
        docs_uploaded = False
        # Map document field names to folder names by category
        doc_fields: dict[str, str] = {}
        if category == 'part':
            doc_fields = {
                'part_qualita': 'qualita',
                'part_3_1_materiale': '3_1_materiale',
                'part_step_tavole': 'step_tavole',
                'part_altro': 'altro',
            }
        elif category == 'assembly':
            doc_fields = {
                'ass_qualita': 'qualita',
                'ass_step_tavole': 'step_tavole',
                'ass_funzionamento': 'funzionamento',
                'ass_istruzioni': 'istruzioni',
                'ass_altro': 'altro',
            }
        else:
            doc_fields = {
                'comm_qualita': 'qualita',
                'comm_ddt_fornitore': 'ddt_fornitore',
                'comm_step_tavole': 'step_tavole',
                'comm_3_1_materiale': '3_1_materiale',
                'comm_altro': 'altro',
            }
        # Persist uploaded documents into ``static/tmp_structures/<type>/<folder>``
        if doc_fields:
            def _safe(n: str) -> str:
                return secure_filename(n) or 'unnamed'
            prod_dir = _safe(st.name)
            base_path = os.path.join(current_app.static_folder, 'tmp_structures', prod_dir)
            for field_name, folder_name in doc_fields.items():
                files = request.files.getlist(field_name) or []
                for f in files:
                    if f and f.filename:
                        filename = secure_filename(f.filename)
                        target_dir = os.path.join(base_path, folder_name)
                        try:
                            os.makedirs(target_dir, exist_ok=True)
                        except Exception:
                            pass
                        try:
                            f.save(os.path.join(target_dir, filename))
                            docs_uploaded = True
                        except Exception:
                            pass

        # Handle image upload if present.  Images are named with a ``st_<id>_`` prefix
        # and stored under ``static/uploads``.
        image_file = request.files.get('image')
        if image_file and _allowed_image(image_file.filename):
            filename = secure_filename(image_file.filename)
            filename = f"st_{st.id}_{filename}"
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            try:
                os.makedirs(upload_dir, exist_ok=True)
            except Exception:
                pass
            try:
                image_file.save(os.path.join(upload_dir, filename))
            except Exception:
                # Image upload is optional; ignore failures
                pass

        # Determine if this POST includes attribute fields beyond documents.
        attribute_keys = [
            'description', 'weight', 'work_phase_id', 'processing_type', 'supplier_id',
            'work_center_id', 'standard_time', 'lead_time_theoretical', 'lead_time_real',
            'processing_cost', 'price_per_unit', 'minimum_order_qty',
            'stock_threshold', 'replenishment_qty'
        ]
        has_attribute_fields = any(k in request.form for k in attribute_keys)

        # When no attribute fields are present (assembly-only documents) return immediately.
        if not has_attribute_fields:
            if docs_uploaded:
                flash('Documenti caricati.', 'success')
            return redirect(url_for('admin.structures'))

        # Update default attributes.  Capture description, raw notes and weight.
        st.default_description = request.form.get('description') or None
        raw_notes_text = request.form.get('notes') or None
        weight_val = request.form.get('weight') or None
        try:
            st.default_weight = float(weight_val) if weight_val else None
        except (ValueError, TypeError):
            st.default_weight = None
        # Inventory defaults: update stock threshold and replenishment quantity
        stock_val = request.form.get('stock_threshold') or None
        repl_val = request.form.get('replenishment_qty') or None
        try:
            st.default_stock_threshold = float(stock_val) if stock_val not in (None, '') else None
        except (ValueError, TypeError):
            st.default_stock_threshold = None
        try:
            st.default_replenishment_qty = float(repl_val) if repl_val not in (None, '') else None
        except (ValueError, TypeError):
            st.default_replenishment_qty = None
        if category == 'part':
            # Populate defaults for parts
            phase_id = request.form.get('work_phase_id') or None
            st.default_work_phase_id = int(phase_id) if phase_id else None
            proc_type = request.form.get('processing_type') or None
            st.default_processing_type = proc_type
            supp_id = request.form.get('supplier_id') or None
            st.default_supplier_id = int(supp_id) if supp_id else None
            center_id = request.form.get('work_center_id') or None
            st.default_work_center_id = int(center_id) if center_id else None
            std_input = request.form.get('standard_time') or None
            std_minutes = None
            try:
                std_hours = float(std_input) if std_input else None
                std_minutes = std_hours * 60.0 if std_hours is not None else None
            except (ValueError, TypeError):
                std_minutes = None
            st.default_standard_time = std_minutes
            def to_minutes(s_val):
                try:
                    return float(s_val) * 1440.0
                except Exception:
                    return None
            st.default_lead_time_theoretical = to_minutes(request.form.get('lead_time_theoretical'))
            st.default_lead_time_real = to_minutes(request.form.get('lead_time_real'))
            proc_cost = request.form.get('processing_cost') or None
            st.default_processing_cost = None
            if proc_cost:
                try:
                    st.default_processing_cost = float(proc_cost)
                except (ValueError, TypeError):
                    st.default_processing_cost = None
            # If no explicit cost provided, compute based on processing type and work centre
            if st.default_processing_cost is None:
                if st.default_processing_type == 'internal' and std_minutes is not None and st.default_work_center_id:
                    wc_obj = WorkCenter.query.get(st.default_work_center_id)
                    if wc_obj and wc_obj.hourly_cost is not None:
                        st.default_processing_cost = (std_minutes / 60.0) * wc_obj.hourly_cost
        elif category == 'commercial':
            supp_id = request.form.get('supplier_id') or None
            st.default_supplier_id = int(supp_id) if supp_id else None
            price_val = request.form.get('price_per_unit') or None
            try:
                st.default_price_per_unit = float(price_val) if price_val else None
            except (ValueError, TypeError):
                st.default_price_per_unit = None
            min_qty = request.form.get('minimum_order_qty') or None
            try:
                st.default_minimum_order_qty = int(min_qty) if min_qty else None
            except (ValueError, TypeError):
                st.default_minimum_order_qty = None
            def to_minutes(s_val):
                try:
                    return float(s_val) * 1440.0
                except Exception:
                    return None
            st.default_lead_time_theoretical = to_minutes(request.form.get('lead_time_theoretical'))
            st.default_lead_time_real = to_minutes(request.form.get('lead_time_real'))
        # Handle additional cycles and compute aggregated processing cost if applicable
        import json
        cycles_json_str = request.form.get('cycles_json') or None
        additional_cycles: list = []
        if cycles_json_str:
            try:
                parsed_cycles = json.loads(cycles_json_str)
                # Clean and filter cycles: drop rows where all values are blank
                # and normalise empty strings to None.  This prevents saving
                # spurious empty cycles when users add and then remove rows.
                if isinstance(parsed_cycles, list):
                    for cyc in parsed_cycles:
                        if not isinstance(cyc, dict):
                            continue
                        if any(cyc.get(key) for key in [
                            'work_phase_id', 'processing_type', 'supplier_id',
                            'work_center_id', 'standard_time', 'lead_time_theoretical',
                            'lead_time_real', 'processing_cost'
                        ]):
                            cleaned = {
                                'work_phase_id': cyc.get('work_phase_id') or None,
                                'processing_type': cyc.get('processing_type') or None,
                                'supplier_id': cyc.get('supplier_id') or None,
                                'work_center_id': cyc.get('work_center_id') or None,
                                'standard_time': cyc.get('standard_time') or None,
                                'lead_time_theoretical': cyc.get('lead_time_theoretical') or None,
                                'lead_time_real': cyc.get('lead_time_real') or None,
                                'processing_cost': cyc.get('processing_cost') or None,
                            }
                            additional_cycles.append(cleaned)
            except Exception:
                additional_cycles = []
        aggregated_cost = None
        proc_cost_input = request.form.get('processing_cost') or None
        if proc_cost_input:
            try:
                aggregated_cost = float(proc_cost_input)
            except (ValueError, TypeError):
                aggregated_cost = None
        if aggregated_cost is None and category == 'part':
            total_cost = 0.0
            # Primary cycle cost for internal processes; convert minutes to hours
            if st.default_processing_type == 'internal' and st.default_standard_time is not None and st.default_work_center_id:
                wc_obj = WorkCenter.query.get(st.default_work_center_id)
                if wc_obj and wc_obj.hourly_cost is not None:
                    try:
                        total_cost += (st.default_standard_time / 60.0) * float(wc_obj.hourly_cost)
                    except Exception:
                        pass
            # Additional cycle costs
            for cycle in additional_cycles:
                try:
                    ctype = cycle.get('processing_type')
                    if ctype == 'internal':
                        st_hours = cycle.get('standard_time')
                        wc_id = cycle.get('work_center_id')
                        if st_hours and wc_id:
                            try:
                                std_h = float(st_hours)
                                wc_obj = WorkCenter.query.get(int(wc_id))
                                if wc_obj and wc_obj.hourly_cost is not None:
                                    total_cost += std_h * float(wc_obj.hourly_cost)
                            except Exception:
                                pass
                    elif ctype == 'external':
                        cost_str = cycle.get('processing_cost')
                        if cost_str:
                            try:
                                total_cost += float(cost_str)
                            except (ValueError, TypeError):
                                pass
                except Exception:
                    pass
            aggregated_cost = total_cost if total_cost > 0 else None
        if aggregated_cost is not None:
            st.default_processing_cost = aggregated_cost
        # Compose default_notes as JSON when cycles exist; otherwise store raw notes
        if additional_cycles:
            try:
                st.default_notes = json.dumps({"notes": raw_notes_text or "", "cycles": additional_cycles})
            except Exception:
                st.default_notes = raw_notes_text
        else:
            st.default_notes = raw_notes_text
        db.session.commit()
        flash('Attributi predefiniti salvati.', 'success')
        return redirect(url_for('admin.structures'))
    # On GET, display a default‑definition page matching the product component editing view.
    # Construct a lightweight 'product' and 'component' object to reuse the products/component_detail template.
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()

    # Create a dummy product for display purposes.  The id is set to 0 to indicate
    # this is not a real product.  The name is derived from the structure type name.
    dummy_product = SimpleNamespace(id=0, name=f"Definizione {st.name}")
    # Build a minimal structure proxy so component_detail.html can access structure.name and structure.type
    structure_proxy = SimpleNamespace(name=st.name, type=st)
    # Build a dummy component with attributes mirroring those on ProductComponent.
    # Use the default values from the StructureType for parts and commercial types.
    dummy_component = SimpleNamespace()
    dummy_component.structure = structure_proxy
    dummy_component.quantity = 1
    dummy_component.description = st.default_description or ''
    dummy_component.notes = st.default_notes or ''
    dummy_component.weight = st.default_weight
    # Processing parameters for parts
    dummy_component.work_phase_id = st.default_work_phase_id
    dummy_component.processing_type = st.default_processing_type
    dummy_component.supplier_id = st.default_supplier_id
    dummy_component.work_center_id = st.default_work_center_id
    dummy_component.standard_time = st.default_standard_time
    dummy_component.lead_time_theoretical = st.default_lead_time_theoretical
    dummy_component.lead_time_real = st.default_lead_time_real
    dummy_component.processing_cost = st.default_processing_cost
    # Commercial parameters
    dummy_component.price_per_unit = st.default_price_per_unit
    dummy_component.minimum_order_qty = st.default_minimum_order_qty

    # Inventory parameters: assign the default stock threshold and replenishment
    # quantity from the structure type.  These values are used to prefill
    # the inventory management fields on the definition form.  Use None when
    # the defaults are not set.
    try:
        dummy_component.stock_threshold = st.default_stock_threshold
    except Exception:
        dummy_component.stock_threshold = None
    try:
        dummy_component.replenishment_qty = st.default_replenishment_qty
    except Exception:
        dummy_component.replenishment_qty = None
    # Image and timestamps default to None for definition pages.  When a
    # default image has been saved for this type, locate it in the uploads
    # folder using a prefix based on the type id and assign it to
    # ``image_filename`` so the template will render it.
    dummy_component.image_filename = None
    try:
        upload_dir = os.path.join(current_app.static_folder, 'uploads')
        prefix = f"st_{st.id}_"
        if os.path.isdir(upload_dir):
            for _fname in os.listdir(upload_dir):
                if _fname.startswith(prefix):
                    dummy_component.image_filename = _fname
                    break
    except Exception:
        dummy_component.image_filename = None
    dummy_component.created_at = None
    dummy_component.updated_at = None
    # Resolve related objects used by the template (e.g. work centre) when available
    if dummy_component.work_center_id:
        dummy_component.work_center = WorkCenter.query.get(dummy_component.work_center_id)
    else:
        dummy_component.work_center = None
    # Compute processing cost on the fly for internal processing if not explicitly set
    if dummy_component.processing_cost is None and dummy_component.processing_type == 'internal' and dummy_component.standard_time is not None and dummy_component.work_center_id:
        wc = dummy_component.work_center
        if wc and wc.hourly_cost is not None:
            # standard_time is stored in minutes; convert to hours
            dummy_component.processing_cost = (dummy_component.standard_time / 60.0) * wc.hourly_cost
    # Render the same template used for product component editing.  No additional cycles are defined for
    # default definitions, so pass an empty list.  Navigation links back to the product will be hidden
    # because created_at is None (see the template for logic).
    # ---------------------------------------------------------------------
    # Gather any documents that have been uploaded for this structure type.  When
    # uploading default documents for a type the files are stored under
    # static/tmp_structures/<type>/<folder>.  Build a mapping of folder
    # identifiers to lists of files (name and relative path) so the template
    # can render links to the existing documents.
    existing_documents: dict[str, list[dict]] = {}
    if category == 'part':
        doc_folders = ['qualita', '3_1_materiale', 'step_tavole', 'altro']
    elif category == 'assembly':
        doc_folders = ['qualita', 'step_tavole', 'funzionamento', 'istruzioni', 'altro']
    else:
        doc_folders = ['qualita', 'ddt_fornitore', 'step_tavole', '3_1_materiale', 'altro']
    def _safe(name: str) -> str:
        return secure_filename(name) or 'unnamed'
    prod_dir = _safe(st.name)
    base_path = os.path.join(current_app.static_folder, 'tmp_structures', prod_dir)
    for folder in doc_folders:
        files: list[dict] = []
        folder_path = os.path.join(base_path, folder)
        try:
            if os.path.isdir(folder_path):
                for fname in os.listdir(folder_path):
                    full_path = os.path.join(folder_path, fname)
                    if os.path.isfile(full_path):
                        rel_path = os.path.join('tmp_structures', prod_dir, folder, fname)
                        files.append({'name': fname, 'path': rel_path})
        except Exception:
            pass
        existing_documents[folder] = files
    doc_label_map: dict[str, str] = {
        'qualita': 'Modulo Cert. qualità',
        '3_1_materiale': '3.1 Materiale',
        'step_tavole': 'Step/tavola',
        'funzionamento': 'Verifica funzionamento',
        'istruzioni': 'Montaggio istruzioni',
        'ddt_fornitore': 'DDT fornitore',
        'altro': 'Altro',
    }
    # Extract a plain text notes value for the template.  The component.notes may
    # contain JSON when cycles are stored, so attempt to parse it and fall back
    # to the raw string.  This mirrors the logic in the product component
    # detail view.  If parsing fails, use the original notes string.
    notes_value = ''
    try:
        import json
        if dummy_component.notes:
            parsed = json.loads(dummy_component.notes)
            if isinstance(parsed, dict) and 'notes' in parsed:
                notes_value = parsed.get('notes', '')
            else:
                notes_value = dummy_component.notes
    except Exception:
        notes_value = dummy_component.notes or ''

    # Parse any saved cycles from the default notes field.  When cycles are
    # present, default_notes is stored as a JSON object with keys
    # ``notes`` and ``cycles``.  Extract the cycles list so that the
    # template can render additional phase rows, and recompute the
    # aggregated cost if not explicitly set.  Also update notes_value
    # accordingly.
    additional_cycles: list = []
    try:
        import json
        parsed_obj = json.loads(dummy_component.notes) if dummy_component.notes else None
        if isinstance(parsed_obj, dict):
            # When cycles are stored, update the notes value from the dict
            additional_cycles = parsed_obj.get('cycles', []) if parsed_obj.get('cycles') else []
            # notes_value has been extracted above; no change needed
    except Exception:
        additional_cycles = []
    # Compute aggregated cost for the dummy component if needed.  The cost
    # may not include additional cycles by default, so recompute when
    # cycles exist and category is 'part'.
    if category == 'part' and additional_cycles:
        total_cost = 0.0
        # Primary cycle cost
        if dummy_component.processing_type == 'internal' and dummy_component.standard_time is not None and dummy_component.work_center_id:
            try:
                wc_obj = WorkCenter.query.get(dummy_component.work_center_id)
                if wc_obj and wc_obj.hourly_cost is not None:
                    total_cost += (dummy_component.standard_time / 60.0) * float(wc_obj.hourly_cost)
            except Exception:
                pass
        # Additional cycles cost
        for cycle in additional_cycles:
            try:
                ctype = cycle.get('processing_type')
                if ctype == 'internal':
                    st_hours = cycle.get('standard_time')
                    wc_id = cycle.get('work_center_id')
                    if st_hours and wc_id:
                        try:
                            std_h = float(st_hours)
                            wc_obj = WorkCenter.query.get(int(wc_id))
                            if wc_obj and wc_obj.hourly_cost is not None:
                                total_cost += std_h * float(wc_obj.hourly_cost)
                        except Exception:
                            pass
                elif ctype == 'external':
                    cost_str = cycle.get('processing_cost')
                    if cost_str:
                        try:
                            total_cost += float(cost_str)
                        except (ValueError, TypeError):
                            pass
            except Exception:
                pass
        dummy_component.processing_cost = total_cost if total_cost > 0 else None
    # Determine flagged documents for this structure type.  Because this page
    # edits defaults at the type level, there may not be a concrete structure
    # identifier to use.  Attempt to retrieve the underlying id from the
    # StructureType (st) when available.  The checklist uses structure
    # identifiers as strings, so default to an empty set when no id is
    # resolvable or no flags are defined.
    try:
        cl_map = load_checklist()
        struct_id = None
        try:
            # Attempt to use the id of the associated structure type as
            # surrogate identifier.  When st.id is present the flagged docs
            # mapping may include an entry for this id; otherwise no docs
            # will be preselected.
            struct_id = st.id
        except Exception:
            struct_id = None
        if struct_id is not None and str(struct_id) in cl_map:
            raw_paths = cl_map.get(str(struct_id), []) or []
            # Normalise and collect paths into a set for quick lookup
            norm_set: set[str] = set()
            for p in raw_paths:
                if not isinstance(p, str):
                    continue
                try:
                    norm = p.replace('\\', '/').strip()
                except Exception:
                    norm = p
                if norm:
                    norm_set.add(norm)
            flagged_docs = norm_set
            # Derive auxiliary sets of filenames and base names (case-insensitive)
            flagged_filenames: set[str] = set()
            flagged_basenames: set[str] = set()
            for path in norm_set:
                parts = path.split('/')
                if parts:
                    fname = parts[-1]
                    fname_lower = str(fname).lower()
                    flagged_filenames.add(fname_lower)
                    base_no_ext = fname_lower.rsplit('.', 1)[0]
                    base_original = base_no_ext.split('_compiled_')[0]
                    if base_original:
                        flagged_basenames.add(base_original)
        else:
            flagged_docs = set()
            flagged_filenames = set()
            flagged_basenames = set()
    except Exception:
        flagged_docs = set()
        flagged_filenames = set()
        flagged_basenames = set()
    return render_template(
        'products/component_detail.html',
        product=dummy_product,
        component=dummy_component,
        category=category,
        suppliers=suppliers,
        work_centers=work_centers,
        work_phases=work_phases,
        additional_cycles=additional_cycles,
        definition_mode=True,
        existing_documents=existing_documents,
        doc_label_map=doc_label_map,
        notes_value=notes_value,
        # Provide the master image filename so the template can display it when the node has no own image.
        master_image_filename=master_image_filename,
        # Pass flagged documents and derived filename/base name sets so the checklist can remain active
        flagged_docs=flagged_docs,
        flagged_filenames=flagged_filenames,
        flagged_basenames=flagged_basenames
    )

@admin_bp.route('/structures/node/<int:node_id>/defaults', methods=['GET', 'POST'])
@login_required
def define_structure_node_defaults(node_id: int):
    """Display and update default attributes for a structure node.

    This view mirrors the product component editing page to assign defaults for a node.
    Changes will be propagated to existing product components referencing this node.
    """
    admin_required()
    # Ensure inventory columns are present before performing queries or updates.
    _ensure_inventory_columns()
    s = Structure.query.get_or_404(node_id)
    # Capture optional return_to and pc_id parameters.  When return_to is present,
    # identifies a URL that the user should be returned to after editing
    # defaults (typically a product detail page with a highlight).  It is
    # propagated through POST redirects and passed to the template so
    # that the "Indietro" button can link back appropriately.
    return_to_param = request.args.get('return_to')
    # pc_id identifies the specific ProductComponent being edited when
    # this view is invoked from a product context.  It allows pre-filling
    # and updating the quantity for only that component.  When absent,
    # quantity updates apply to all components referencing this node.
    pc_id_param = request.args.get('pc_id')
    # Determine category from flags
    if s.flag_assembly:
        category = 'assembly'
    elif s.flag_part:
        category = 'part'
    else:
        category = 'commercial'
    if request.method == 'POST':
        """
        Handle saving of documents, images and default attributes for structure nodes.

        This POST handler has been refactored to update the global ComponentMaster
        associated with this node rather than storing most attributes directly
        on the Structure itself.  Each component code (e.g. "P001-025-001B")
        corresponds to a single master record.  When editing defaults for a
        node, the master is updated and changes automatically propagate to
        all structures and product components sharing the same code.
        """
        # Ensure there is a ComponentMaster for this structure before updating
        try:
            ensure_component_master_for_structure(s)
        except Exception:
            pass
        master = getattr(s, 'component_master', None)
        # Decide document fields based on category
        doc_fields: dict[str, str] = {}
        if category == 'part':
            doc_fields = {
                'part_qualita': 'qualita',
                'part_3_1_materiale': '3_1_materiale',
                'part_step_tavole': 'step_tavole',
                'part_altro': 'altro',
            }
        elif category == 'assembly':
            doc_fields = {
                'ass_qualita': 'qualita',
                'ass_step_tavole': 'step_tavole',
                'ass_funzionamento': 'funzionamento',
                'ass_istruzioni': 'istruzioni',
                'ass_altro': 'altro',
            }
        else:  # commercial
            doc_fields = {
                'comm_qualita': 'qualita',
                'comm_ddt_fornitore': 'ddt_fornitore',
                'comm_step_tavole': 'step_tavole',
                'comm_3_1_materiale': '3_1_materiale',
                'comm_altro': 'altro',
            }
        # -------------------------------------------------------------------
        # Save uploaded documents.  When a master exists, documents are stored
        # under ``static/tmp_components/<code>/<folder>``.  Otherwise they fall
        # back to the legacy ``tmp_structures`` location.  The docs_uploaded
        # flag tracks whether any file was saved to show a success message.
        docs_uploaded = False
        if doc_fields:
            def _safe_name(name: str) -> str:
                return secure_filename(name) or 'unnamed'
            # Determine base path.  Prefer the master path for a global code.
            base_path: str | None = None
            if master:
                try:
                    base_path = os.path.join(
                        current_app.static_folder,
                        'tmp_components',
                        _safe_name(master.code)
                    )
                except Exception:
                    base_path = None
            if not base_path:
                # Fallback to tmp_structures path (legacy) if master missing
                type_dir = _safe_name(s.type.name)
                struct_parts: list[str] = []
                def _collect_path(node_obj):
                    if node_obj.parent:
                        _collect_path(node_obj.parent)
                    struct_parts.append(_safe_name(node_obj.name))
                _collect_path(s)
                base_path = os.path.join(current_app.static_folder, 'tmp_structures', type_dir, *struct_parts)
            for field_name, folder_name in doc_fields.items():
                files = request.files.getlist(field_name) or []
                for f in files:
                    if f and f.filename:
                        filename = secure_filename(f.filename)
                        target_dir = os.path.join(base_path, folder_name)
                        try:
                            os.makedirs(target_dir, exist_ok=True)
                        except Exception:
                            pass
                        try:
                            f.save(os.path.join(target_dir, filename))
                            docs_uploaded = True
                        except Exception:
                            pass
        # -------------------------------------------------------------------
        # Save uploaded image.  If a master exists the image is stored with
        # prefix ``cm_<id>_`` so it is shared globally.  Otherwise fall back
        # to the legacy ``sn_<structure id>_`` prefix.
        def _allowed_image(filename: str) -> bool:
            return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
        image_file = request.files.get('image')
        if image_file and _allowed_image(image_file.filename):
            filename = secure_filename(image_file.filename)
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_dir, exist_ok=True)
            if master:
                dest_name = f"cm_{master.id}_{filename}"
            else:
                dest_name = f"sn_{s.id}_{filename}"
            try:
                image_file.save(os.path.join(upload_dir, dest_name))
            except Exception:
                pass
        # -------------------------------------------------------------------
        # Save compatible revisions selections on POST.  The "compatible_revisions"
        # field may be submitted either via the main save form or via the
        # separate revision form.  Capture all values with getlist() so
        # multiple checkboxes or multi‑select options are handled.  If at
        # least one compatible revision is provided, join them with commas
        # and store on the structure.  When none are provided, set
        # compatible_revisions to None to clear any previous selections.
        try:
            # Collect all submitted compatibility values.  A multi‑select and
            # hidden inputs submit the currently selected revision(s).  Use
            # getlist() to gather multiple checkbox/select values.  When
            # only a single comma‑separated string is posted (e.g. from
            # a multi‑select in some browsers), split it on commas.  The
            # current revision itself should not be stored, so filter it out.
            compat_vals = request.form.getlist('compatible_revisions') or []
            if not compat_vals:
                raw_val = request.form.get('compatible_revisions')
                if raw_val:
                    # Split comma‑separated string into individual values
                    compat_vals = [v.strip() for v in raw_val.split(',') if v.strip()]
            # Determine presence of the field explicitly to distinguish
            # between no submissions (retain existing value) and empty
            # submission (clear value).  request.form lists keys for
            # submitted inputs.
            if 'compatible_revisions' in request.form or 'current_revision' in request.form or compat_vals:
                # Build a list of selected previous revisions, preserving
                # order, and skip the current revision (s.revision_label).
                selected_prev: list[str] = []
                try:
                    for letter in s.revision_letters:
                        if letter == s.revision_label:
                            continue
                        if letter in compat_vals and letter not in selected_prev:
                            selected_prev.append(letter)
                except Exception:
                    # Fallback: if revision_letters is unavailable, use the raw values
                    for val in compat_vals:
                        if val != s.revision_label and val not in selected_prev:
                            selected_prev.append(val)
                s.compatible_revisions = ','.join(selected_prev) if selected_prev else None
                # Persist the compatibility selections immediately.  Without
                # committing here the change could be lost if subsequent logic
                # returns early (e.g. documents-only upload) or an exception
                # occurs before the next commit.  Refresh the in-memory
                # structure instance so template rendering after redirect
                # reflects the updated list.
                try:
                    db.session.commit()
                    db.session.refresh(s)
                except Exception:
                    # If commit or refresh fails, rollback to a clean state.
                    db.session.rollback()
        except Exception:
            # Ignore errors; compatibility is optional
            pass
        # -------------------------------------------------------------------
        # Determine whether the form contains fields other than the revision
        # compatibility controls.  Only treat a request as an attribute update
        # when at least one recognised attribute key (e.g. description, weight,
        # supplier_id, etc.) is present.  Compatibility fields
        # ("compatible_revisions" and "current_revision") are deliberately
        # excluded here so that selecting or deselecting compatible revisions
        # does not inadvertently trigger a full attribute update cycle (which
        # would reset unspecified fields on the master).  When no attribute
        # fields are present (i.e. the user only toggled compatibility or
        # uploaded documents), skip the subsequent attribute update logic and
        # redirect early after flashing any document upload messages.
        attribute_keys = [
            'description', 'weight', 'work_phase_id', 'processing_type', 'supplier_id',
            'work_center_id', 'standard_time', 'lead_time_theoretical', 'lead_time_real',
            'processing_cost', 'price_per_unit', 'minimum_order_qty',
            # Inventory management fields: include stock threshold and replenishment quantity so
            # that a POST containing only these values will still be processed as an attribute
            # update rather than being treated as a documents-only submission.
            'stock_threshold', 'replenishment_qty',
            'quantity', 'lot_management', 'is_sellable', 'guiding_part'
        ]
        has_other_attributes = any(k in request.form for k in attribute_keys)
        # If there are no other attribute fields besides compatibility controls,
        # return early after processing compatibility and/or document uploads.
        if not has_other_attributes:
            if docs_uploaded:
                flash('Documenti caricati.', 'success')
            # When only compatibility or documents were updated, avoid
            # executing the remainder of the attribute update logic.  Redirect
            # back to the appropriate page.  If the request originated from
            # a product page, remain on this node definition page so the user
            # can continue editing.  Otherwise go back to the structures list.
            if return_to_param:
                return redirect(url_for('admin.define_structure_node_defaults', node_id=node_id, return_to=return_to_param))
            return redirect(url_for('admin.structures'))
        # -------------------------------------------------------------------
        # Extract sellable and guiding flags from the form.  The presence of
        # these checkboxes indicates a True value; absence means False.  This
        # applies to all categories.  We update the master when present,
        # otherwise fall back to the structure itself.  Additionally,
        # propagate these flags to all product components referencing this
        # structure to keep values consistent across products.
        sellable_flag = True if request.form.get('is_sellable') else False
        guiding_flag = True if request.form.get('guiding_part') else False
        try:
            if master:
                master.is_sellable = sellable_flag
                master.guiding_part = guiding_flag
            else:
                s.is_sellable = sellable_flag
                s.guiding_part = guiding_flag
            # Propagate to all related product components.  If a specific pc_id
            # is provided, update only that component; otherwise update
            # components filtered by product id or all components.
            comps_to_update: list[ProductComponent] = []
            if pc_id_param:
                try:
                    pc_int = int(pc_id_param)
                    comp = ProductComponent.query.get(pc_int)
                    if comp:
                        comps_to_update.append(comp)
                except Exception:
                    comps_to_update = comps_to_update
            elif return_to_param:
                try:
                    import re
                    m_prod = re.search(r"/products/(\d+)", return_to_param)
                    if m_prod:
                        pid = int(m_prod.group(1))
                        comp = ProductComponent.query.filter_by(product_id=pid, structure_id=s.id).first()
                        if comp:
                            comps_to_update.append(comp)
                except Exception:
                    pass
            # If no specific component context, update all components referencing this node
            if not comps_to_update:
                comps_to_update = ProductComponent.query.filter_by(structure_id=s.id).all()
            for _c in comps_to_update:
                try:
                    _c.is_sellable = sellable_flag
                    _c.guiding_part = guiding_flag
                except Exception:
                    pass
            db.session.commit()
        except Exception:
            # Ignore any errors; flags are optional
            db.session.rollback()
        # -------------------------------------------------------------------
        # Handle quantity updates for this node.  When a quantity field is
        # provided, update the quantity on all product components referencing
        # this structure.  Only positive integers are considered valid.
        qty_val = request.form.get('quantity') or None
        if qty_val:
            try:
                qty_int = int(qty_val)
                if qty_int > 0:
                    # Determine which components to update.  If a specific pc_id was
                    # provided, update only that component.  Else, if return_to
                    # contains a product id, update the component belonging to that
                    # product.  Otherwise update all components referencing this
                    # structure.
                    updated = False
                    if pc_id_param:
                        try:
                            pcid_int = int(pc_id_param)
                            comp = ProductComponent.query.get(pcid_int)
                            if comp:
                                comp.quantity = qty_int
                                db.session.commit()
                                updated = True
                        except Exception:
                            pass
                    elif return_to_param:
                        try:
                            import re
                            m_prod = re.search(r"/products/(\d+)", return_to_param)
                            if m_prod:
                                prod_id = int(m_prod.group(1))
                                comp = ProductComponent.query.filter_by(product_id=prod_id, structure_id=s.id).first()
                                if comp:
                                    comp.quantity = qty_int
                                    db.session.commit()
                                    updated = True
                        except Exception:
                            pass
                    if not updated:
                        comps = ProductComponent.query.filter_by(structure_id=s.id).all()
                        for comp in comps:
                            comp.quantity = qty_int
                        db.session.commit()
            except Exception:
                # Ignore invalid quantity input
                pass
        # -------------------------------------------------------------------
        # At this point we have attribute fields to update.  Update the master
        # record with the canonical information.  Fall back to updating the
        # structure directly only when a master is not present.
        raw_notes_text = request.form.get('notes') or None
        # Helper: convert days to minutes
        def _to_minutes(days_val: str | None) -> float | None:
            try:
                return float(days_val) * 1440.0 if days_val not in (None, '') else None
            except Exception:
                return None
        if master:
            # Update description and weight
            master.description = request.form.get('description') or None
            weight_val = request.form.get('weight') or None
            try:
                master.weight = float(weight_val) if weight_val else None
            except (ValueError, TypeError):
                master.weight = None
            # Reset numeric/relational fields to avoid stale values when cleared
            master.processing_type = None
            master.work_phase_id = None
            master.supplier_id = None
            master.work_center_id = None
            master.standard_time = None
            master.lead_time_theoretical = None
            master.lead_time_real = None
            master.processing_cost = None
            master.price_per_unit = None
            master.minimum_order_qty = None

            # Inventory management: update stock threshold and replenishment quantity on the master.
            # Read values from the form and convert to floats when provided.  If the
            # input is empty or invalid, store None to indicate the field is unset.
            stock_val = request.form.get('stock_threshold') or None
            repl_val = request.form.get('replenishment_qty') or None
            try:
                master.stock_threshold = float(stock_val) if stock_val not in (None, '') else None
            except (ValueError, TypeError):
                master.stock_threshold = None
            try:
                master.replenishment_qty = float(repl_val) if repl_val not in (None, '') else None
            except (ValueError, TypeError):
                master.replenishment_qty = None
            # Handle part-specific fields and cycles
            if category == 'part':
                phase_id = request.form.get('work_phase_id') or None
                master.work_phase_id = int(phase_id) if phase_id else None
                master.processing_type = request.form.get('processing_type') or None
                supp_id = request.form.get('supplier_id') or None
                # Safely convert supplier identifier to integer if possible
                if supp_id:
                    try:
                        master.supplier_id = int(supp_id)
                    except (ValueError, TypeError):
                        master.supplier_id = None
                else:
                    master.supplier_id = None
                center_id = request.form.get('work_center_id') or None
                master.work_center_id = int(center_id) if center_id else None
                # Standard time (hours -> minutes)
                std_input = request.form.get('standard_time') or None
                try:
                    std_hours = float(std_input) if std_input else None
                except (ValueError, TypeError):
                    std_hours = None
                master.standard_time = std_hours * 60.0 if std_hours is not None else None
                master.lead_time_theoretical = _to_minutes(request.form.get('lead_time_theoretical'))
                master.lead_time_real = _to_minutes(request.form.get('lead_time_real'))
                # Parse additional cycles (list of dicts) from cycles_json hidden field
                import json
                cycles_json_str = request.form.get('cycles_json') or None
                additional_cycles: list = []
                if cycles_json_str:
                    try:
                        parsed_cycles = json.loads(cycles_json_str)
                        if isinstance(parsed_cycles, list):
                            for cyc in parsed_cycles:
                                if not isinstance(cyc, dict):
                                    continue
                                if any(cyc.get(key) for key in [
                                    'work_phase_id', 'processing_type', 'supplier_id',
                                    'work_center_id', 'standard_time', 'lead_time_theoretical',
                                    'lead_time_real', 'processing_cost'
                                ]):
                                    cleaned = {
                                        'work_phase_id': cyc.get('work_phase_id') or None,
                                        'processing_type': cyc.get('processing_type') or None,
                                        'supplier_id': cyc.get('supplier_id') or None,
                                        'work_center_id': cyc.get('work_center_id') or None,
                                        'standard_time': cyc.get('standard_time') or None,
                                        'lead_time_theoretical': cyc.get('lead_time_theoretical') or None,
                                        'lead_time_real': cyc.get('lead_time_real') or None,
                                        'processing_cost': cyc.get('processing_cost') or None,
                                    }
                                    additional_cycles.append(cleaned)
                    except Exception:
                        additional_cycles = []
                # Fallback: build cycles from repeated form fields if cycles_json missing
                if not additional_cycles:
                    try:
                        phases_all = request.form.getlist('work_phase_id')
                        procs_all = request.form.getlist('processing_type')
                        suppliers_all = request.form.getlist('supplier_id')
                        centers_all = request.form.getlist('work_center_id')
                        stds_all = request.form.getlist('standard_time')
                        lt_theo_all = request.form.getlist('lead_time_theoretical')
                        lt_real_all = request.form.getlist('lead_time_real')
                        total_cycles = len(procs_all)
                        for idx in range(1, total_cycles):
                            cyc_phase = phases_all[idx] if idx < len(phases_all) else ''
                            cyc_proc = procs_all[idx] if idx < len(procs_all) else ''
                            cyc_supplier = suppliers_all[idx] if idx < len(suppliers_all) else ''
                            cyc_center = centers_all[idx] if idx < len(centers_all) else ''
                            cyc_std = stds_all[idx] if idx < len(stds_all) else ''
                            cyc_lt_theo = lt_theo_all[idx] if idx < len(lt_theo_all) else ''
                            cyc_lt_real = lt_real_all[idx] if idx < len(lt_real_all) else ''
                            if any([cyc_phase, cyc_proc, cyc_supplier, cyc_center, cyc_std, cyc_lt_theo, cyc_lt_real]):
                                additional_cycles.append({
                                    'work_phase_id': cyc_phase or None,
                                    'processing_type': cyc_proc or None,
                                    'supplier_id': cyc_supplier or None,
                                    'work_center_id': cyc_center or None,
                                    'standard_time': cyc_std or None,
                                    'lead_time_theoretical': cyc_lt_theo or None,
                                    'lead_time_real': cyc_lt_real or None,
                                    'processing_cost': None,
                                })
                    except Exception:
                        additional_cycles = additional_cycles
                # Determine aggregated processing cost.  Prefer client-provided value
                aggregated_cost: float | None = None
                proc_cost_input = request.form.get('processing_cost') or None
                if proc_cost_input:
                    try:
                        aggregated_cost = float(proc_cost_input)
                    except (ValueError, TypeError):
                        aggregated_cost = None
                if aggregated_cost is None:
                    total_cost = 0.0
                    # cost for primary cycle (internal only)
                    if master.processing_type == 'internal' and std_hours is not None and master.work_center_id:
                        wc_obj = WorkCenter.query.get(master.work_center_id)
                        if wc_obj and wc_obj.hourly_cost is not None:
                            try:
                                total_cost += std_hours * float(wc_obj.hourly_cost)
                            except Exception:
                                pass
                    # cost for additional cycles
                    for cycle in additional_cycles:
                        try:
                            ctype = cycle.get('processing_type')
                            if ctype == 'internal':
                                st_val = cycle.get('standard_time')
                                wc_id = cycle.get('work_center_id')
                                if st_val and wc_id:
                                    try:
                                        st_float = float(st_val)
                                        wc_obj = WorkCenter.query.get(int(wc_id))
                                        if wc_obj and wc_obj.hourly_cost is not None:
                                            total_cost += st_float * float(wc_obj.hourly_cost)
                                    except Exception:
                                        pass
                            elif ctype == 'external':
                                cost_str = cycle.get('processing_cost')
                                if cost_str:
                                    try:
                                        total_cost += float(cost_str)
                                    except (ValueError, TypeError):
                                        pass
                        except Exception:
                            pass
                    aggregated_cost = total_cost if total_cost > 0 else None
                master.processing_cost = aggregated_cost
                # Compose notes JSON if cycles exist; otherwise store raw notes
                if additional_cycles:
                    try:
                        master.notes = json.dumps({"notes": raw_notes_text or "", "cycles": additional_cycles})
                    except Exception:
                        master.notes = raw_notes_text
                else:
                    master.notes = raw_notes_text
            elif category == 'commercial':
                supp_id = request.form.get('supplier_id') or None
                # Safely convert supplier identifier to integer if possible
                if supp_id:
                    try:
                        master.supplier_id = int(supp_id)
                    except (ValueError, TypeError):
                        master.supplier_id = None
                else:
                    master.supplier_id = None
                price_val = request.form.get('price_per_unit') or None
                try:
                    master.price_per_unit = float(price_val) if price_val else None
                except (ValueError, TypeError):
                    master.price_per_unit = None
                min_qty = request.form.get('minimum_order_qty') or None
                try:
                    master.minimum_order_qty = int(min_qty) if min_qty else None
                except (ValueError, TypeError):
                    master.minimum_order_qty = None
                master.lead_time_theoretical = _to_minutes(request.form.get('lead_time_theoretical'))
                master.lead_time_real = _to_minutes(request.form.get('lead_time_real'))
                # Compose notes JSON with optional lot_management flag
                lot_flag = True if request.form.get('lot_management') else False
                notes_input = raw_notes_text
                notes_dict = {}
                # Always include a notes key when lot management is enabled
                # to avoid storing a JSON object with only ``lot_management``.
                if notes_input or lot_flag:
                    notes_dict['notes'] = notes_input or ''
                if lot_flag:
                    notes_dict['lot_management'] = True
                if notes_dict:
                    try:
                        import json as _json
                        master.notes = _json.dumps(notes_dict)
                    except Exception:
                        master.notes = notes_input
                else:
                    master.notes = None
            else:  # assemblies
                # Assemblies only store description and notes
                master.notes = raw_notes_text
            # Persist master updates
            db.session.commit()
            # Optionally mirror master fields onto structure for backward compatibility
            try:
                s.description = master.description
                s.weight = master.weight
                s.processing_type = master.processing_type
                s.work_phase_id = master.work_phase_id
                s.supplier_id = master.supplier_id
                s.work_center_id = master.work_center_id
                s.standard_time = master.standard_time
                s.lead_time_theoretical = master.lead_time_theoretical
                s.lead_time_real = master.lead_time_real
                s.processing_cost = master.processing_cost
                s.price_per_unit = master.price_per_unit
                s.minimum_order_qty = master.minimum_order_qty
                s.notes = master.notes
                db.session.commit()
            except Exception:
                db.session.rollback()
        else:
            # No master available; update the structure as a fallback (legacy behaviour)
            # Update description, weight and notes
            s.description = request.form.get('description') or None
            weight_val = request.form.get('weight') or None
            try:
                s.weight = float(weight_val) if weight_val else None
            except (ValueError, TypeError):
                s.weight = None
            s.notes = raw_notes_text

            # Inventory management: update stock threshold and replenishment quantity on the structure
            stock_val = request.form.get('stock_threshold') or None
            repl_val = request.form.get('replenishment_qty') or None
            try:
                s.stock_threshold = float(stock_val) if stock_val not in (None, '') else None
            except (ValueError, TypeError):
                s.stock_threshold = None
            try:
                s.replenishment_qty = float(repl_val) if repl_val not in (None, '') else None
            except (ValueError, TypeError):
                s.replenishment_qty = None
            # Continue to update category-specific fields on structure
            if category == 'part':
                phase_id = request.form.get('work_phase_id') or None
                s.work_phase_id = int(phase_id) if phase_id else None
                proc_type = request.form.get('processing_type') or None
                s.processing_type = proc_type
                supp_id = request.form.get('supplier_id') or None
                s.supplier_id = int(supp_id) if supp_id else None
                center_id = request.form.get('work_center_id') or None
                s.work_center_id = int(center_id) if center_id else None
                std_input = request.form.get('standard_time') or None
                try:
                    std_hours = float(std_input) if std_input else None
                except (ValueError, TypeError):
                    std_hours = None
                s.standard_time = std_hours * 60.0 if std_hours is not None else None
                s.lead_time_theoretical = _to_minutes(request.form.get('lead_time_theoretical'))
                s.lead_time_real = _to_minutes(request.form.get('lead_time_real'))
                try:
                    s.processing_cost = float(request.form.get('processing_cost')) if request.form.get('processing_cost') else None
                except (ValueError, TypeError):
                    s.processing_cost = None
            elif category == 'commercial':
                supp_id = request.form.get('supplier_id') or None
                s.supplier_id = int(supp_id) if supp_id else None
                try:
                    s.price_per_unit = float(request.form.get('price_per_unit')) if request.form.get('price_per_unit') else None
                except (ValueError, TypeError):
                    s.price_per_unit = None
                try:
                    s.minimum_order_qty = int(request.form.get('minimum_order_qty')) if request.form.get('minimum_order_qty') else None
                except (ValueError, TypeError):
                    s.minimum_order_qty = None
                s.lead_time_theoretical = _to_minutes(request.form.get('lead_time_theoretical'))
                s.lead_time_real = _to_minutes(request.form.get('lead_time_real'))
                # Compose notes JSON with optional lot_management flag
                lot_flag = True if request.form.get('lot_management') else False
                notes_input = raw_notes_text
                notes_dict = {}
                # Always include a notes key when lot management is enabled
                # to avoid storing a JSON object with only ``lot_management``.
                if notes_input or lot_flag:
                    notes_dict['notes'] = notes_input or ''
                if lot_flag:
                    notes_dict['lot_management'] = True
                if notes_dict:
                    try:
                        import json as _json
                        s.notes = _json.dumps(notes_dict)
                    except Exception:
                        s.notes = notes_input
                else:
                    s.notes = None
            db.session.commit()
        # End of attribute handling
        flash('Attributi del nodo salvati.', 'success')

        # -----------------------------------------------------------------
        # Propagate inventory fields (stock threshold and replenishment quantity)
        # to the product components referencing this structure.  When editing
        # defaults from a product context, update only the targeted
        # component identified by ``pc_id_param`` or by the product id in
        # ``return_to_param``.  Otherwise update all components that
        # reference this node.  Values are taken from the master when
        # available, falling back to the structure.  Errors during
        # propagation are silently ignored to avoid interrupting the
        # editing flow.
        try:
            # Determine values to propagate
            stock_to_propagate = None
            repl_to_propagate = None
            if master:
                stock_to_propagate = master.stock_threshold
                repl_to_propagate = master.replenishment_qty
            else:
                stock_to_propagate = s.stock_threshold
                repl_to_propagate = s.replenishment_qty
            comps_to_update: list[ProductComponent] = []
            if pc_id_param:
                try:
                    pc_int = int(pc_id_param)
                    comp = ProductComponent.query.get(pc_int)
                    if comp:
                        comps_to_update.append(comp)
                except Exception:
                    comps_to_update = comps_to_update
            elif return_to_param:
                try:
                    import re as _re
                    m_prod = _re.search(r"/products/(\d+)", return_to_param)
                    if m_prod:
                        pid = int(m_prod.group(1))
                        comp = ProductComponent.query.filter_by(product_id=pid, structure_id=s.id).first()
                        if comp:
                            comps_to_update.append(comp)
                except Exception:
                    pass
            # If no specific component context, update all components referencing this node
            if not comps_to_update:
                comps_to_update = ProductComponent.query.filter_by(structure_id=s.id).all()
            for _c in comps_to_update:
                try:
                    _c.stock_threshold = stock_to_propagate
                    _c.replenishment_qty = repl_to_propagate
                except Exception:
                    pass
            db.session.commit()
        except Exception:
            db.session.rollback()
        # -----------------------------------------------------------------
        # Propagate updated master attributes to all structures sharing the
        # same component code.  When editing defaults for one node, other
        # nodes with the same code (across different parent structures or
        # products) should reflect the canonical attributes such as
        # description, weight and processing parameters.  Update these
        # structures to mirror the master values.  Skip the current node
        # because it has already been updated above.
        try:
            # Determine flags for the current structure to match only nodes of the
            # same category (assembly/part/commercial) when updating.  This
            # prevents mixing attributes across categories.
            flag_ass = s.flag_assembly
            flag_part = s.flag_part
            flag_comm = s.flag_commercial
            # Fetch all structures with the same name and category
            others = (Structure.query
                      .filter(Structure.name == s.name,
                              Structure.flag_assembly == flag_ass,
                              Structure.flag_part == flag_part,
                              Structure.flag_commercial == flag_comm)
                      .filter(Structure.id != s.id)
                      .all())
            for other in others:
                # Mirror master attributes onto the other structure
                other.description = master.description
                other.weight = master.weight
                other.processing_type = master.processing_type
                other.work_phase_id = master.work_phase_id
                other.supplier_id = master.supplier_id
                other.work_center_id = master.work_center_id
                other.standard_time = master.standard_time
                other.lead_time_theoretical = master.lead_time_theoretical
                other.lead_time_real = master.lead_time_real
                other.processing_cost = master.processing_cost
                other.price_per_unit = master.price_per_unit
                other.minimum_order_qty = master.minimum_order_qty
                other.notes = master.notes
            if others:
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
        except Exception:
            # Silently ignore propagation errors
            pass
        # After saving attributes and propagating them, remain on the definition
        # page if the user arrived from a product context.  Otherwise return
        # to the structures overview.  Preserve the return_to parameter on the
        # redirect.
        if return_to_param:
            return redirect(url_for('admin.define_structure_node_defaults', node_id=node_id, return_to=return_to_param))
        return redirect(url_for('admin.structures'))
    # On GET, present a page identical to the product component editing view.  Construct
    # a dummy product and component so the existing template can be reused.  This
    # allows administrators to define defaults in a familiar environment.
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()

    dummy_product = SimpleNamespace(id=0, name=f"Definizione {s.name}")
    # For node defaults use the actual structure as the component's structure.  The
    # dummy component mirrors the fields of the structure and, when a component
    # master exists for this node, it populates missing or blank values from
    # the master.  This ensures that when editing defaults for a node whose
    # code already exists globally (e.g. "P001-025-001B") the form shows the
    # canonical information rather than empty fields.
    dummy_component = SimpleNamespace()
    dummy_component.structure = s
    # Default quantity is 1.  When editing from a product context (return_to
    # provided), attempt to prefill the quantity from the corresponding
    # ProductComponent for that product.  Otherwise leave at 1 because
    # default definitions do not track a quantity.
    dummy_component.quantity = 1
    # Prefer values from the component master when available
    master = getattr(s, 'component_master', None)
    # Description and notes
    if master and getattr(master, 'description', None):
        dummy_component.description = master.description
    else:
        dummy_component.description = s.description or ''
    if master and getattr(master, 'notes', None):
        dummy_component.notes = master.notes
    else:
        dummy_component.notes = s.notes or ''
    # Weight
    if master and master.weight is not None:
        dummy_component.weight = master.weight
    else:
        dummy_component.weight = s.weight
    # Processing parameters for parts
    dummy_component.work_phase_id = None
    dummy_component.processing_type = None
    dummy_component.supplier_id = None
    dummy_component.work_center_id = None
    dummy_component.standard_time = None
    dummy_component.lead_time_theoretical = None
    dummy_component.lead_time_real = None
    dummy_component.processing_cost = None
    # Prefer master values when present; fallback to node values
    if master:
        dummy_component.work_phase_id = master.work_phase_id if master.work_phase_id is not None else s.work_phase_id
        dummy_component.processing_type = master.processing_type if master.processing_type is not None else s.processing_type
        dummy_component.supplier_id = master.supplier_id if master.supplier_id is not None else s.supplier_id
        dummy_component.work_center_id = master.work_center_id if master.work_center_id is not None else s.work_center_id
        dummy_component.standard_time = master.standard_time if master.standard_time is not None else s.standard_time
        dummy_component.lead_time_theoretical = master.lead_time_theoretical if master.lead_time_theoretical is not None else s.lead_time_theoretical
        dummy_component.lead_time_real = master.lead_time_real if master.lead_time_real is not None else s.lead_time_real
        dummy_component.processing_cost = master.processing_cost if master.processing_cost is not None else s.processing_cost
    else:
        dummy_component.work_phase_id = s.work_phase_id
        dummy_component.processing_type = s.processing_type
        dummy_component.supplier_id = s.supplier_id
        dummy_component.work_center_id = s.work_center_id
        dummy_component.standard_time = s.standard_time
        dummy_component.lead_time_theoretical = s.lead_time_theoretical
        dummy_component.lead_time_real = s.lead_time_real
        dummy_component.processing_cost = s.processing_cost
    # Commercial parameters
    dummy_component.price_per_unit = None
    dummy_component.minimum_order_qty = None
    if master:
        dummy_component.price_per_unit = master.price_per_unit if master.price_per_unit is not None else s.price_per_unit
        dummy_component.minimum_order_qty = master.minimum_order_qty if master.minimum_order_qty is not None else s.minimum_order_qty
    else:
        dummy_component.price_per_unit = s.price_per_unit
        dummy_component.minimum_order_qty = s.minimum_order_qty

    # -----------------------------------------------------------------
    # Inventory parameters: prefill the stock threshold and replenishment quantity
    # on the dummy component.  These values are used for display on the
    # definition page and reflect the current settings on the master or the
    # structure itself.  When a ComponentMaster exists for this node, use
    # its values; otherwise fall back to the structure-level values.  If
    # neither provides a value, leave as None.
    try:
        if master and getattr(master, 'stock_threshold', None) is not None:
            dummy_component.stock_threshold = master.stock_threshold
        else:
            dummy_component.stock_threshold = s.stock_threshold
    except Exception:
        dummy_component.stock_threshold = getattr(s, 'stock_threshold', None)
    try:
        if master and getattr(master, 'replenishment_qty', None) is not None:
            dummy_component.replenishment_qty = master.replenishment_qty
        else:
            dummy_component.replenishment_qty = s.replenishment_qty
    except Exception:
        dummy_component.replenishment_qty = getattr(s, 'replenishment_qty', None)

    # ---------------------------------------------------------------------
    # Prefill sellable and guiding flags.  When a master exists, prefer
    # its values; otherwise fall back to the structure values.  If neither
    # provides a value, default to False.
    try:
        if master and getattr(master, 'is_sellable', None) is not None:
            dummy_component.is_sellable = bool(master.is_sellable)
        elif getattr(s, 'is_sellable', None) is not None:
            dummy_component.is_sellable = bool(s.is_sellable)
        else:
            dummy_component.is_sellable = False
    except Exception:
        dummy_component.is_sellable = False
    try:
        if master and getattr(master, 'guiding_part', None) is not None:
            dummy_component.guiding_part = bool(master.guiding_part)
        elif getattr(s, 'guiding_part', None) is not None:
            dummy_component.guiding_part = bool(s.guiding_part)
        else:
            dummy_component.guiding_part = False
    except Exception:
        dummy_component.guiding_part = False

    # ---------------------------------------------------------------------
    # Prefill quantity from a product component when editing defaults
    # via a product context.  If a pc_id parameter is provided, look up
    # that specific ProductComponent and use its quantity.  Otherwise, if
    # return_to_param exists, parse the product id from the URL and
    # retrieve the component by product_id and structure_id.  Only
    # positive quantities are considered; if none found, keep the default.
    if pc_id_param:
        try:
            pcid_int = int(pc_id_param)
            comp = ProductComponent.query.get(pcid_int)
            if comp and comp.quantity:
                dummy_component.quantity = comp.quantity
        except Exception:
            pass
    elif return_to_param:
        try:
            import re
            match = re.search(r"/products/(\d+)", return_to_param)
            if match:
                prod_id = int(match.group(1))
                comp = ProductComponent.query.filter_by(product_id=prod_id, structure_id=s.id).first()
                if comp and comp.quantity:
                    dummy_component.quantity = comp.quantity
        except Exception:
            pass

    # Determine if there is a master image for this node to display in the
    # component detail template.  A master image has the prefix "cm_<id>_" in
    # the static uploads directory.  If found, store the filename for the
    # template.  If no master or no file is found, leave as None.
    master_image_filename = None
    try:
        m = getattr(s, 'component_master', None)
        if m:
            upload_dir = os.path.join(current_app.static_folder, 'uploads')
            if os.path.isdir(upload_dir):
                prefix = f"cm_{m.id}_"
                for f in os.listdir(upload_dir):
                    if f.startswith(prefix):
                        master_image_filename = f
                        break
    except Exception:
        master_image_filename = None
    # Image and timestamps are None for default definitions
    # Determine if a default image has been saved for this node.  Images
    # uploaded via the defaults form are stored under static/uploads with a
    # prefix ``sn_<node_id>_``.  Search for the first matching file and
    # assign its name to image_filename.  When no image exists, set to None
    # so the template displays "Nessuna immagine".
    dummy_component.image_filename = None
    try:
        upload_dir = os.path.join(current_app.static_folder, 'uploads')
        prefix = f"sn_{s.id}_"
        if os.path.isdir(upload_dir):
            for _fname in os.listdir(upload_dir):
                if _fname.startswith(prefix):
                    dummy_component.image_filename = _fname
                    break
    except Exception:
        dummy_component.image_filename = None
    dummy_component.created_at = None
    dummy_component.updated_at = None
    # Resolve related work centre if available
    if dummy_component.work_center_id:
        dummy_component.work_center = WorkCenter.query.get(dummy_component.work_center_id)
    else:
        dummy_component.work_center = None
    # Compute processing cost for internal processing if missing
    if dummy_component.processing_cost is None and dummy_component.processing_type == 'internal' and dummy_component.standard_time is not None and dummy_component.work_center_id:
        wc = dummy_component.work_center
        if wc and wc.hourly_cost is not None:
            dummy_component.processing_cost = (dummy_component.standard_time / 60.0) * wc.hourly_cost
    # ---------------------------------------------------------------------
    # Extract any additional cycles from the notes field.  Similar to product
    # components, the notes string may contain JSON with keys "notes" and
    # "cycles".  If cycles are present, assign them to additional_cycles and
    # update dummy_component.notes to contain only the plain notes for display.
    additional_cycles: list = []
    if category == 'part':
        try:
            import json
            if dummy_component.notes:
                parsed = json.loads(dummy_component.notes)
                if isinstance(parsed, dict) and 'cycles' in parsed:
                    cycles_data = parsed.get('cycles')
                    if isinstance(cycles_data, list):
                        additional_cycles = cycles_data
                    dummy_component.notes = parsed.get('notes', '')
        except Exception:
            additional_cycles = []
        # If processing cost is missing, recompute based on primary and additional cycles
        # to display a meaningful value.  Use dummy_component values for primary
        # cycle and the extracted cycles for additional cycles.
        if dummy_component.processing_cost is None:
            total_cost = 0.0
            # Primary cost: internal only
            if dummy_component.processing_type == 'internal' and dummy_component.standard_time is not None and dummy_component.work_center_id:
                wc_obj = WorkCenter.query.get(dummy_component.work_center_id)
                if wc_obj and wc_obj.hourly_cost is not None:
                    try:
                        total_cost += (dummy_component.standard_time / 60.0) * float(wc_obj.hourly_cost)
                    except Exception:
                        pass
            # Additional cycles costs
            for cycle in additional_cycles:
                try:
                    ctype = cycle.get('processing_type')
                    if ctype == 'internal':
                        st = cycle.get('standard_time')
                        wc_id = cycle.get('work_center_id')
                        if st and wc_id:
                            try:
                                st_float = float(st)
                                wc_obj = WorkCenter.query.get(int(wc_id))
                                if wc_obj and wc_obj.hourly_cost is not None:
                                    total_cost += st_float * float(wc_obj.hourly_cost)
                            except Exception:
                                pass
                    elif ctype == 'external':
                        cost_str = cycle.get('processing_cost')
                        if cost_str:
                            try:
                                total_cost += float(cost_str)
                            except (ValueError, TypeError):
                                pass
                except Exception:
                    pass
            dummy_component.processing_cost = total_cost if total_cost > 0 else None
    # Build a mapping of existing documents for this node.  Uploaded documents
    # for structure nodes are stored under static/tmp_structures/<type>/<node path>/<folder>.
    existing_documents: dict[str, list[dict]] = {}
    # Determine the document folders for the category
    if category == 'part':
        doc_folders = ['qualita', '3_1_materiale', 'step_tavole', 'altro']
    elif category == 'assembly':
        doc_folders = ['qualita', 'step_tavole', 'funzionamento', 'istruzioni', 'altro']
    else:
        doc_folders = ['qualita', 'ddt_fornitore', 'step_tavole', '3_1_materiale', 'altro']
    # Build the base path for documents.  Prefer the master path when available.
    def _safe(name: str) -> str:
        return secure_filename(name) or 'unnamed'
    master = getattr(s, 'component_master', None)
    base_path: str | None = None
    if master:
        try:
            base_path = os.path.join(current_app.static_folder, 'tmp_components', _safe(master.code))
        except Exception:
            base_path = None
    if not base_path:
        # Fallback to tmp_structures path (legacy)
        type_dir = _safe(s.type.name)
        struct_parts: list[str] = []
        def _collect(node):
            if node.parent:
                _collect(node.parent)
            struct_parts.append(_safe(node.name))
        _collect(s)
        base_path = os.path.join(current_app.static_folder, 'tmp_structures', type_dir, *struct_parts)
    # For each folder gather files relative to appropriate base path
    for folder in doc_folders:
        files: list[dict] = []
        folder_path = os.path.join(base_path, folder)
        try:
            if os.path.isdir(folder_path):
                for fname in os.listdir(folder_path):
                    full_path = os.path.join(folder_path, fname)
                    if os.path.isfile(full_path):
                        # Build relative path for links.  Use different base directories
                        # depending on whether the path is tmp_components or tmp_structures.
                        if 'tmp_components' in base_path:
                            rel_path = os.path.join('tmp_components', _safe(master.code) if master else '', folder, fname)
                        else:
                            # Legacy structure path relative part
                            type_dir = _safe(s.type.name)
                            struct_parts: list[str] = []
                            def _collect_rel(node):
                                if node.parent:
                                    _collect_rel(node.parent)
                                struct_parts.append(_safe(node.name))
                            _collect_rel(s)
                            rel_path = os.path.join('tmp_structures', type_dir, *struct_parts, folder, fname)
                        files.append({'name': fname, 'path': rel_path})
        except Exception:
            pass
        existing_documents[folder] = files
    doc_label_map: dict[str, str] = {
        'qualita': 'Modulo Cert. qualità',
        '3_1_materiale': '3.1 Materiale',
        'step_tavole': 'Step/tavola',
        'funzionamento': 'Verifica funzionamento',
        'istruzioni': 'Montaggio istruzioni',
        'ddt_fornitore': 'DDT fornitore',
        'altro': 'Altro',
    }
    # Extract a plain notes value for the template.  The notes field may
    # contain JSON when cycles are stored, similar to ProductComponent notes.
    # Attempt to parse JSON and extract the 'notes' key; otherwise use the
    # raw string.  If parsing fails fall back to the original value.
    notes_value = ''
    try:
        import json
        if dummy_component.notes:
            parsed = json.loads(dummy_component.notes)
            if isinstance(parsed, dict) and 'notes' in parsed:
                notes_value = parsed.get('notes', '')
            else:
                notes_value = dummy_component.notes
    except Exception:
        notes_value = dummy_component.notes or ''

    # Render the editing template for the structure node.  Pass through
    # return_to_param so the template can render a back link when editing
    # defaults from a product context.
    # -------------------------------------------------------------------
    # Compute the aggregate weight of all descendant parts when editing
    # defaults for an assembly.  The ``component_detail`` template expects
    # a ``parts_weight`` variable to be defined for assemblies so that it
    # can display the total weight of child parts.  When this value is
    # missing Jinja may throw an error.  We replicate the logic from the
    # product blueprint but omit quantity multipliers (assume quantity=1)
    # because quantities are defined at the ProductComponent level and
    # unavailable when editing structure defaults.  For non-assembly nodes
    # parts_weight remains None.
    parts_weight: float | None = None
    if category == 'assembly':
        total_weight = 0.0
        def _accumulate_weights(node: Structure):
            """Recursively sum the weights of all descendant parts and
            commercial components under the given structure.  Assemblies
            contribute the aggregated weight of their children while parts
            contribute their own weight.  Quantities are assumed to be 1
            because product-specific quantities are unavailable in this
            context."""
            nonlocal total_weight
            try:
                # Traverse each child structure
                for child in node.children:
                    # Determine the weight for this child by preferring the
                    # component master weight and falling back to the
                    # structure's own weight.  Skip if neither is defined.
                    child_weight = None
                    try:
                        cm = getattr(child, 'component_master', None)
                        if cm and cm.weight is not None:
                            child_weight = cm.weight
                        elif child.weight is not None:
                            child_weight = child.weight
                    except Exception:
                        # If any attribute lookup fails, fallback to structure weight
                        try:
                            child_weight = child.weight
                        except Exception:
                            child_weight = None
                    # If the child is an assembly, recurse to accumulate the
                    # weights of its descendants rather than using its own
                    # weight directly.  This ensures nested assemblies are
                    # accounted for correctly.
                    if child.flag_assembly:
                        _accumulate_weights(child)
                    else:
                        # Only add the weight if defined.  Quantities are
                        # treated as 1 because we do not have product-specific
                        # quantity information in this context.
                        if child_weight is not None:
                            try:
                                total_weight += float(child_weight)
                            except Exception:
                                pass
            except Exception:
                pass
        try:
            _accumulate_weights(s)
            parts_weight = total_weight
        except Exception:
            # If any error occurs leave parts_weight as None
            parts_weight = None
    # -------------------------------------------------------------------
    # Determine whether the lot management flag should be preselected.  When
    # editing defaults for commercial parts, the notes field may contain
    # JSON with a "lot_management" boolean.  Attempt to parse this flag
    # from the dummy component's notes and, if absent, from the master or
    # structure notes.  Default to False.
    lot_management_flag = False
    if category == 'commercial':
        try:
            import json as _json
            # First look at dummy_component.notes which may be a JSON string or plain text
            if getattr(dummy_component, 'notes', None):
                try:
                    parsed_lot = _json.loads(dummy_component.notes)
                    if isinstance(parsed_lot, dict) and 'lot_management' in parsed_lot:
                        lot_management_flag = bool(parsed_lot.get('lot_management'))
                except Exception:
                    # notes may be plain text; ignore
                    pass
            # If not found on the dummy component, check master notes
            if not lot_management_flag:
                m = getattr(s, 'component_master', None)
                if m and m.notes:
                    try:
                        parsed_m = _json.loads(m.notes)
                        if isinstance(parsed_m, dict) and 'lot_management' in parsed_m:
                            lot_management_flag = bool(parsed_m.get('lot_management'))
                    except Exception:
                        pass
            # Fallback to structure notes if still not found
            if not lot_management_flag and s.notes:
                try:
                    parsed_s = _json.loads(s.notes)
                    if isinstance(parsed_s, dict) and 'lot_management' in parsed_s:
                        lot_management_flag = bool(parsed_s.get('lot_management'))
                except Exception:
                    pass
        except Exception:
            lot_management_flag = False
    # -------------------------------------------------------------------
    # Determine flagged documents for this structure node.  Use the node's id to
    # look up any document paths recorded in the checklist and derive sets
    # of filenames and base names (both lower-case) so the UI can match
    # documents by name even when paths change due to renaming or revision.
    try:
        cl_map = load_checklist()
        struct_id = None
        try:
            struct_id = s.id
        except Exception:
            struct_id = None
        if struct_id is not None and str(struct_id) in cl_map:
            raw_paths = cl_map.get(str(struct_id), []) or []
            norm_set: set[str] = set()
            for p in raw_paths:
                if not isinstance(p, str):
                    continue
                try:
                    norm = p.replace('\\', '/').strip()
                except Exception:
                    norm = p
                if norm:
                    norm_set.add(norm)
            flagged_docs = norm_set
            # Build filename and base name sets for flexible matching
            flagged_filenames: set[str] = set()
            flagged_basenames: set[str] = set()
            for path in norm_set:
                parts = path.split('/')
                if parts:
                    fname = parts[-1]
                    fname_lower = str(fname).lower()
                    flagged_filenames.add(fname_lower)
                    base_no_ext = fname_lower.rsplit('.', 1)[0]
                    base_original = base_no_ext.split('_compiled_')[0]
                    if base_original:
                        flagged_basenames.add(base_original)
        else:
            flagged_docs = set()
            flagged_filenames = set()
            flagged_basenames = set()
    except Exception:
        flagged_docs = set()
        flagged_filenames = set()
        flagged_basenames = set()
    return render_template(
        'products/component_detail.html',
        product=dummy_product,
        component=dummy_component,
        category=category,
        suppliers=suppliers,
        work_centers=work_centers,
        work_phases=work_phases,
        additional_cycles=additional_cycles,
        definition_mode=True,
        existing_documents=existing_documents,
        doc_label_map=doc_label_map,
        notes_value=notes_value,
        # Provide both the return_to parameter (for back navigation) and
        # master_image_filename so the template can display the image saved
        # on the ComponentMaster when no per-node image exists.
        return_to=return_to_param,
        master_image_filename=master_image_filename,
        # Pass the computed parts_weight so the template can avoid
        # undefined variable errors and display the aggregated weight when
        # editing assembly defaults.  For non-assembly nodes this will
        # remain None.
        parts_weight=parts_weight,
        # Pass the lot management flag to the template so the checkbox is
        # preselected when the current node has lot management enabled.
        lot_management=lot_management_flag,
        # Include flagged documents and derived filename sets so the checklist
        # checkboxes remain selected when matching by file name or base name.
        flagged_docs=flagged_docs,
        flagged_filenames=flagged_filenames,
        flagged_basenames=flagged_basenames
    )

# Nuovi endpoint per la gestione degli utenti


@admin_bp.route('/users')
@login_required
def users():
    """Elenca tutti gli utenti registrati."""
    admin_required()
    users = User.query.order_by(User.username.asc()).all()
    return render_template('admin/users.html', users=users)


@admin_bp.route('/users/new', methods=['GET', 'POST'])
@login_required
def new_user():
    """Crea un nuovo utente."""
    admin_required()
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        role = request.form.get('role', 'user') or 'user'
        if not username or not password or not email:
            flash('Username, email e password sono obbligatorie.', 'danger')
        elif User.query.filter_by(username=username).first():
            flash('Esiste già un utente con questo username.', 'warning')
        elif User.query.filter_by(email=email).first():
            flash('Esiste già un utente con questa email.', 'warning')
        else:
            u = User(username=username, email=email, role=role)
            u.set_password(password)
            db.session.add(u)
            db.session.commit()
            flash('Utente creato.', 'success')
            return redirect(url_for('admin.users'))
    return render_template('admin/new_user.html')


@admin_bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id: int):
    """Aggiorna username, ruolo o password di un utente esistente."""
    admin_required()
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        email = request.form.get('email', '').strip().lower()
        role = request.form.get('role', 'user') or 'user'
        password = request.form.get('password', '')

        if not username or not email:
            flash('Username ed email sono obbligatori.', 'danger')
        elif User.query.filter(User.username == username, User.id != user.id).first():
            flash('Esiste già un utente con questo username.', 'warning')
        elif User.query.filter(User.email == email, User.id != user.id).first():
            flash('Esiste già un utente con questa email.', 'warning')
        else:
            user.username = username
            user.email = email
            user.role = role
            if password:
                user.set_password(password)
            db.session.commit()
            flash('Utente aggiornato.', 'success')
            return redirect(url_for('admin.users'))

    return render_template('admin/edit_user.html', user=user)


@admin_bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id: int):
    """Elimina un utente esistente (non può eliminare se stesso)."""
    admin_required()
    # Prevent deletion of current user
    if current_user.id == user_id:
        flash('Non puoi eliminare il tuo stesso account.', 'warning')
        return redirect(url_for('admin.users'))
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('Utente eliminato.', 'success')
    return redirect(url_for('admin.users'))


# -----------------------------------------------------------------------------
# Dizionario: gestione fornitori, centri di lavoro, fasi di lavorazione e costi
# Questo endpoint mostra le liste delle voci di dizionario e consente di
# aggiungerne di nuove.  L'eliminazione o modifica possono essere implementate
# in seguito.

@admin_bp.route('/dictionary', methods=['GET', 'POST'])
@login_required
def dictionary():
    admin_required()
    # Ensure the ``hourly_cost`` column exists on the ``work_centers`` table.  When
    # upgrading from earlier versions of the application the database may not
    # contain this column yet, causing an OperationalError when querying.
    # We perform a runtime check and add the column if it is missing.  This
    # operation is idempotent on SQLite.  On other databases explicit
    # migrations are preferred but are not available in this simplified setup.
    try:
        # Use a direct connection to inspect the table schema via PRAGMA.
        from sqlalchemy import text
        conn = db.engine.connect()
        res = conn.execute(text('PRAGMA table_info(work_centers)')).fetchall()
        col_names = [row[1] for row in res]  # second column contains the name
        if 'hourly_cost' not in col_names:
            # Add the column with a generic float type.  SQLite will accept
            # multiple ADD COLUMN statements; only executed once when missing.
            conn.execute(text('ALTER TABLE work_centers ADD COLUMN hourly_cost FLOAT'))
        conn.close()
    except Exception:
        # If inspection fails, silently ignore; missing column will be detected
        # later and raise an error.  Logging could be added here if needed.
        pass
    if request.method == 'POST':
        category = request.form.get('category')
        # handle each category separately; for material cost we expect two fields
        if category == 'supplier':
            name = request.form.get('name', '').strip()
            if name and not Supplier.query.filter_by(name=name).first():
                db.session.add(Supplier(name=name))
                db.session.commit()
                flash('Fornitore aggiunto.', 'success')
        elif category == 'work_center':
            # Aggiungi un nuovo centro di lavoro con nome e costo orario facoltativo
            name = request.form.get('name', '').strip()
            cost_val = request.form.get('cost', '').strip()
            # Converte il costo in un numero se presente; se non valido lascia None
            try:
                cost = float(cost_val) if cost_val else None
            except (ValueError, TypeError):
                cost = None
            if name and not WorkCenter.query.filter_by(name=name).first():
                db.session.add(WorkCenter(name=name, hourly_cost=cost))
                db.session.commit()
                flash('Centro di lavoro aggiunto.', 'success')
        elif category == 'work_phase':
            name = request.form.get('name', '').strip()
            if name and not WorkPhase.query.filter_by(name=name).first():
                db.session.add(WorkPhase(name=name))
                db.session.commit()
                flash('Fase di lavorazione aggiunta.', 'success')
        elif category == 'material_cost':
            material = request.form.get('material', '').strip()
            cost_val = request.form.get('cost', '').strip()
            if material:
                try:
                    cost = float(cost_val) if cost_val else None
                except ValueError:
                    cost = None
                db.session.add(MaterialCost(material=material, cost_eur=cost))
                db.session.commit()
                flash('Costo materiale aggiunto.', 'success')
        return redirect(url_for('admin.dictionary'))
    # GET: mostra elenco voci
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()
    material_costs = MaterialCost.query.order_by(MaterialCost.material.asc()).all()
    return render_template('admin/dictionary.html',
                           suppliers=suppliers,
                           work_centers=work_centers,
                           work_phases=work_phases,
                           material_costs=material_costs)


# -----------------------------------------------------------------------------
# Dizionario: editing e eliminazione voci
#
# Per ciascuna voce del dizionario (fornitore, centro di lavoro, fase di
# lavorazione e costo materiale) definiamo due endpoint: uno per modificare la
# voce esistente e uno per eliminarla.  Tutti gli endpoint richiedono che
# l'utente sia autenticato come amministratore.  Le operazioni di eliminazione
# includono un controllo delle dipendenze per evitare di rimuovere voci ancora
# utilizzate da componenti o prodotti.

@admin_bp.route('/dictionary/supplier/<int:supplier_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_supplier(supplier_id: int):
    """Modifica un fornitore esistente.

    Visualizza un modulo precompilato con il nome del fornitore; al submit
    aggiorna il record se il nuovo nome non è vuoto né duplicato.
    """
    admin_required()
    supplier = Supplier.query.get_or_404(supplier_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Il nome è obbligatorio.', 'danger')
        else:
            # Assicurarsi che non esista un altro fornitore con lo stesso nome
            existing = (
                Supplier.query
                .filter(Supplier.name == name, Supplier.id != supplier.id)
                .first()
            )
            if existing:
                flash('Esiste già un fornitore con questo nome.', 'warning')
            else:
                supplier.name = name
                db.session.commit()
                flash('Fornitore aggiornato.', 'success')
                return redirect(url_for('admin.dictionary'))
    return render_template('admin/edit_supplier.html', supplier=supplier)


@admin_bp.route('/dictionary/supplier/<int:supplier_id>/delete', methods=['POST'])
@login_required
def delete_supplier(supplier_id: int):
    """Elimina un fornitore se non è utilizzato da alcun componente.

    Il metodo POST viene chiamato dal pulsante "Elimina" nella vista
    dizionario.  Se esistono riferimenti in ProductComponent, l'eliminazione
    viene bloccata e viene mostrato un messaggio d'errore.
    """
    admin_required()
    supplier = Supplier.query.get_or_404(supplier_id)
    # Verifica se il fornitore è referenziato da qualche componente
    referenced = ProductComponent.query.filter_by(supplier_id=supplier.id).first()
    if referenced:
        flash('Impossibile eliminare il fornitore: esistono componenti che lo utilizzano.', 'danger')
    else:
        db.session.delete(supplier)
        db.session.commit()
        flash('Fornitore eliminato.', 'success')
    return redirect(url_for('admin.dictionary'))


@admin_bp.route('/dictionary/work_center/<int:center_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_work_center(center_id: int):
    """Modifica un centro di lavoro esistente.

    Consente di aggiornare il nome e il costo orario associato.  Il costo
    viene convertito in un numero con passo 0.0001; valori non numerici
    vengono ignorati e trattati come None.
    """
    admin_required()
    center = WorkCenter.query.get_or_404(center_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        cost_val = request.form.get('cost', '').strip()
        try:
            cost = float(cost_val) if cost_val else None
        except (ValueError, TypeError):
            cost = None
        if not name:
            flash('Il nome è obbligatorio.', 'danger')
        else:
            # Controlla unicità del nome
            existing = (
                WorkCenter.query
                .filter(WorkCenter.name == name, WorkCenter.id != center.id)
                .first()
            )
            if existing:
                flash('Esiste già un centro di lavoro con questo nome.', 'warning')
            else:
                center.name = name
                center.hourly_cost = cost
                db.session.commit()
                flash('Centro di lavoro aggiornato.', 'success')
                return redirect(url_for('admin.dictionary'))
    return render_template('admin/edit_work_center.html', center=center)


@admin_bp.route('/dictionary/work_center/<int:center_id>/delete', methods=['POST'])
@login_required
def delete_work_center(center_id: int):
    """Elimina un centro di lavoro se non è referenziato da componenti.

    Prima di cancellare il centro di lavoro si verifica se esistono componenti
    che lo utilizzano tramite la relazione ProductComponent.work_center_id.
    """
    admin_required()
    center = WorkCenter.query.get_or_404(center_id)
    referenced = ProductComponent.query.filter_by(work_center_id=center.id).first()
    if referenced:
        flash('Impossibile eliminare il centro di lavoro: esistono componenti che lo utilizzano.', 'danger')
    else:
        db.session.delete(center)
        db.session.commit()
        flash('Centro di lavoro eliminato.', 'success')
    return redirect(url_for('admin.dictionary'))


@admin_bp.route('/dictionary/work_phase/<int:phase_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_work_phase(phase_id: int):
    """Modifica una fase di lavorazione esistente.

    Permette di cambiare il nome della fase mantenendo la sua chiave primaria.
    """
    admin_required()
    phase = WorkPhase.query.get_or_404(phase_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Il nome è obbligatorio.', 'danger')
        else:
            existing = (
                WorkPhase.query
                .filter(WorkPhase.name == name, WorkPhase.id != phase.id)
                .first()
            )
            if existing:
                flash('Esiste già una fase di lavorazione con questo nome.', 'warning')
            else:
                phase.name = name
                db.session.commit()
                flash('Fase di lavorazione aggiornata.', 'success')
                return redirect(url_for('admin.dictionary'))
    return render_template('admin/edit_work_phase.html', phase=phase)


@admin_bp.route('/dictionary/work_phase/<int:phase_id>/delete', methods=['POST'])
@login_required
def delete_work_phase(phase_id: int):
    """Elimina una fase di lavorazione se non è utilizzata da componenti.

    L'eliminazione viene impedita se esistono componenti associati tramite
    ProductComponent.work_phase_id.
    """
    admin_required()
    phase = WorkPhase.query.get_or_404(phase_id)
    referenced = ProductComponent.query.filter_by(work_phase_id=phase.id).first()
    if referenced:
        flash('Impossibile eliminare la fase di lavorazione: esistono componenti che la utilizzano.', 'danger')
    else:
        db.session.delete(phase)
        db.session.commit()
        flash('Fase di lavorazione eliminata.', 'success')
    return redirect(url_for('admin.dictionary'))


@admin_bp.route('/dictionary/material_cost/<int:mc_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_material_cost(mc_id: int):
    """Modifica un costo materiale esistente.

    Consente di aggiornare il nome del materiale e il costo per chilogrammo.  Il
    costo è opzionale; se non specificato resta None.
    """
    admin_required()
    mc = MaterialCost.query.get_or_404(mc_id)
    if request.method == 'POST':
        material = request.form.get('material', '').strip()
        cost_val = request.form.get('cost', '').strip()
        try:
            cost = float(cost_val) if cost_val else None
        except (ValueError, TypeError):
            cost = None
        if not material:
            flash('Il materiale è obbligatorio.', 'danger')
        else:
            # Non esistono vincoli univoci sul materiale; eventuali
            # duplicati sono consentiti ma possono essere gestiti a livello
            # applicativo se necessario.
            mc.material = material
            mc.cost_eur = cost
            db.session.commit()
            flash('Costo materiale aggiornato.', 'success')
            return redirect(url_for('admin.dictionary'))
    return render_template('admin/edit_material_cost.html', mc=mc)


@admin_bp.route('/dictionary/material_cost/<int:mc_id>/delete', methods=['POST'])
@login_required
def delete_material_cost(mc_id: int):
    """Elimina una voce di costo materiale.

    Attualmente non vengono effettuati controlli sulle dipendenze poiché i
    costi materiali non sono referenziati direttamente da altre tabelle.
    """
    admin_required()
    mc = MaterialCost.query.get_or_404(mc_id)
    db.session.delete(mc)
    db.session.commit()
    flash('Costo materiale eliminato.', 'success')
    return redirect(url_for('admin.dictionary'))

# Placeholder route for importing a BOM CSV file.  This view simply renders
# a form with a file input.  The actual import logic will be implemented later.
@admin_bp.route('/import-structure', methods=['GET'])
@admin_bp.route('/import-structure', methods=['GET', 'POST'])
@login_required
def import_structure():
    """
    Importa una distinta base da un file CSV o Excel e crea le strutture, i nodi e il prodotto
    corrispondente nel database.  L'utente può specificare un nome per il prodotto o, se
    assente, verrà dedotto dal nome del file (prima del punto).  La tabella importata deve
    contenere le colonne 'Num. \nArticolo', 'Num. parte', 'Descrizione', 'Quantità',
    'Peso [g]', 'Materiale' e 'Costo[€]'.  Le righe verranno interpretate secondo il livello
    gerarchico (ad esempio "1", "1.2", ecc.) e saranno classificate come assieme, parte
    o parte commerciale in base al prefisso del campo "Num. parte" ('A' → assieme,
    'P' → parte, altrimenti commerciale).
    """
    admin_required()
    if request.method == 'POST':
        product_name = request.form.get('product_name', '').strip()
        file = request.files.get('bom_file')
        if not file or file.filename == '':
            flash('Seleziona un file da importare.', 'warning')
            return redirect(url_for('admin.import_structure'))
        filename = secure_filename(file.filename)
        # Deduce product name from filename if not provided
        if not product_name:
            # Use part before first space or dot as product name
            base = os.path.splitext(filename)[0]
            product_name = base.split()[0]
        # Read the uploaded file without relying on pandas. Use csv for .csv files
        # and openpyxl for Excel files. Expect a header row with the required fields.
        ext = os.path.splitext(filename)[1].lower()
        rows = []
        try:
            if ext == '.csv' or ext == '.txt':
                import csv
                file.stream.seek(0)
                # decode bytes to text
                decoded = file.stream.read().decode('utf-8-sig', errors='ignore').splitlines()
                reader = csv.reader(decoded)
                header = next(reader)
                # Normalize header names
                # Normalize header names: trim whitespace and replace newlines with spaces
                normalized = []
                for h in header:
                    # Replace newlines/carriage returns with spaces and collapse multiple spaces to a single space
                    temp = h.replace('\r', ' ').replace('\n', ' ')
                    normalized.append(' '.join(temp.split()))
                for row in reader:
                    if len(row) == 0:
                        continue
                    rows.append(dict(zip(normalized, row)))
            elif ext in ('.xls', '.xlsx'):
                # Parse Excel files using openpyxl so we can inspect cell styles (e.g. font colours)
                try:
                    import openpyxl
                except ImportError:
                    flash('La libreria openpyxl non è installata. Impossibile importare il file Excel.', 'danger')
                    return redirect(url_for('admin.import_structure'))
                # Reset stream position and load workbook without stripping styles
                file.stream.seek(0)
                wb = openpyxl.load_workbook(file, data_only=True)
                sheet = wb.active
                # Read the header row (first row) as cell objects so we can normalise names
                header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1)))
                header = []
                for cell in header_cells:
                    val = cell.value
                    val_str = str(val).strip() if val is not None else ''
                    # Normalize header names: replace carriage returns/newlines with spaces and collapse multiple spaces
                    temp = val_str.replace('\r', ' ').replace('\n', ' ')
                    header.append(' '.join(temp.split()))
                # Determine index of "Num. parte" column for colour classification
                try:
                    num_parte_idx = header.index('Num. parte')
                except ValueError:
                    num_parte_idx = None
                # Helper to convert openpyxl colour to an RGB hex string
                def _rgb_from_color(color_obj):
                    """Return a 6‑character RGB string (e.g. 'FF0000') from an openpyxl colour object.
                    Returns None if the colour cannot be determined."""
                    if color_obj is None:
                        return None
                    # When the colour is specified directly via RGB
                    if getattr(color_obj, 'type', None) == 'rgb' and color_obj.rgb:
                        rgb8 = color_obj.rgb  # e.g. 'FFFF0000' (includes alpha channel)
                        # Use the last 6 chars for RGB to ignore alpha if present
                        return rgb8[-6:].upper()
                    # When the colour references the workbook theme, use the theme and tint to compute the colour
                    if getattr(color_obj, 'type', None) == 'theme' and hasattr(color_obj, 'theme'):
                        try:
                            theme_idx = color_obj.theme
                            tint = getattr(color_obj, 'tint', 0) or 0
                            # Map of theme indices to base colours extracted from the workbook theme (see xl/theme/theme1.xml)
                            # The sequence is: dark1, light1, dark2, light2, accent1‑accent6, hyperlink, followed hyperlink
                            theme_palette = ['000000', 'FFFFFF', '0E2841', 'E8E8E8',
                                             '156082', 'E97132', '196B24', '0F9ED5',
                                             'A02B93', '4EA72E', '467886', '96607D']
                            if 0 <= theme_idx < len(theme_palette):
                                base_hex = theme_palette[theme_idx]
                                # Convert to RGB components
                                r = int(base_hex[0:2], 16)
                                g = int(base_hex[2:4], 16)
                                b = int(base_hex[4:6], 16)
                                if tint < 0:
                                    # Darken the colour: multiply each channel
                                    r = int(r * (1 + tint))
                                    g = int(g * (1 + tint))
                                    b = int(b * (1 + tint))
                                else:
                                    # Lighten the colour: linear interpolation towards white
                                    r = int((255 - r) * tint + r)
                                    g = int((255 - g) * tint + g)
                                    b = int((255 - b) * tint + b)
                                # Clamp values and format back into hex
                                r = max(0, min(255, r))
                                g = max(0, min(255, g))
                                b = max(0, min(255, b))
                                return f"{r:02X}{g:02X}{b:02X}"
                        except Exception:
                            # Fall back to None if conversion fails
                            return None
                    # Other colour types (e.g. indexed) are not handled; return None
                    return None
                # Helper to map an RGB colour to a category (assembly, part or commercial)
                def _category_from_rgb(rgb_str):
                    """Determine category based on dominance of red or green components.
                    Returns 'assembly' for predominately red hues, 'commercial' for predominately green hues,
                    'part' for dark/neutral colours, and None if rgb_str is None."""
                    if not rgb_str:
                        return None
                    try:
                        r = int(rgb_str[0:2], 16)
                        g = int(rgb_str[2:4], 16)
                        b = int(rgb_str[4:6], 16)
                    except Exception:
                        return None
                    # Determine if the colour is red‑dominant or green‑dominant
                    # Avoid divisions by using ratio thresholds
                    # Consider a colour red‑dominant if red is significantly greater than green and blue
                    if r > g * 1.3 and r > b * 1.3:
                        return 'assembly'
                    # Consider a colour green‑dominant if green is significantly greater than red and blue
                    if g > r * 1.3 and g > b * 1.3:
                        return 'commercial'
                    # Otherwise classify as part
                    return 'part'
                # Iterate over subsequent rows as cell objects so we can extract font colours
                for row in sheet.iter_rows(min_row=2):
                    # Skip completely blank rows
                    if all((c.value is None or str(c.value).strip() == '') for c in row):
                        continue
                    row_dict: dict[str, str] = {}
                    for i, cell in enumerate(row[:len(header)]):
                        key = header[i]
                        val = cell.value
                        row_dict[key] = str(val).strip() if val is not None else ''
                    # Determine category from colour of the "Num. parte" cell, if present
                    color_category = None
                    if num_parte_idx is not None and num_parte_idx < len(row):
                        color_obj = row[num_parte_idx].font.color
                        rgb = _rgb_from_color(color_obj)
                        color_category = _category_from_rgb(rgb)
                    # Store the colour‑derived category in a private key for later use
                    row_dict['_color_category'] = color_category
                    rows.append(row_dict)
            else:
                flash('Formato file non supportato. Usa CSV o XLSX.', 'danger')
                return redirect(url_for('admin.import_structure'))
        except Exception as e:
            flash(f'Errore nella lettura del file: {e}', 'danger')
            return redirect(url_for('admin.import_structure'))

        # Required column names (case sensitive after normalization)
        # Expected column names after normalization (newlines replaced by spaces)
        required_cols = ['Num. Articolo', 'Num. parte', 'Descrizione', 'Quantità', 'Peso [g]', 'Materiale', 'Costo[€]']
        # Normalize keys for rows: replace any carriage return/newline in keys with a space and trim
        normalized_rows = []
        for rd in rows:
            nrd = {}
            for k, v in rd.items():
                if isinstance(k, str):
                    temp_key = k.replace('\r', ' ').replace('\n', ' ')
                    nk = ' '.join(temp_key.split())
                else:
                    nk = k
                nrd[nk] = v
            normalized_rows.append(nrd)
        rows = normalized_rows
        missing_header = [c for c in required_cols if all(c not in row for row in rows)]
        if missing_header:
            flash(f'Colonne mancanti nel file: {", ".join(missing_header)}', 'danger')
            return redirect(url_for('admin.import_structure'))
        # Create or get product
        product = Product.query.filter_by(name=product_name).first()
        if not product:
            product = Product(name=product_name, description=f'Distinta importata da {filename}')
            db.session.add(product)
            db.session.flush()
        # Create a structure type for the import if it doesn't exist
        type_name = f'Distinta {product_name}'
        stype = StructureType.query.filter_by(name=type_name).first()
        if not stype:
            stype = StructureType(name=type_name, description=f'Tipologia importata da {filename}')
            db.session.add(stype)
            db.session.flush()
        # Mapping for numbering to structures
        struct_map = {}
        for rd in rows:
            num = str(rd.get('Num. Articolo', '')).strip()
            if not num:
                continue
            code = str(rd.get('Num. parte', '')).strip()
            # Determine quantity
            qty_val = rd.get('Quantità', '')
            try:
                qty = int(float(qty_val)) if qty_val != '' else 1
            except (ValueError, TypeError):
                qty = 1
            # Weight kg
            weight_str = rd.get('Peso [g]', '')
            try:
                weight_kg = float(weight_str) / 1000.0 if weight_str not in ('', None) else None
            except (ValueError, TypeError):
                weight_kg = None
            # Material and cost
            material = rd.get('Materiale', '').strip() or None
            cost_str = rd.get('Costo[€]', '')
            try:
                cost = float(cost_str) if cost_str not in ('', None) else None
            except (ValueError, TypeError):
                cost = None
            # Extract description for the component
            desc = rd.get('Descrizione', '').strip()
            # Determine category: prefer classification based on the colour of the "Num. parte" cell.
            # If no colour‑based category is present (e.g. in CSV files or cells without a specific colour),
            # fall back to the prefix logic (A → assieme, P → parte, others → commerciale).
            is_ass = False
            is_part = False
            is_comm = False
            color_cat = rd.get('_color_category')
            if color_cat == 'assembly':
                is_ass = True
            elif color_cat == 'part':
                is_part = True
            elif color_cat == 'commercial':
                is_comm = True
            else:
                # Fallback: use prefix of the code to classify
                prefix = code[0].upper() if code else ''
                is_ass = prefix == 'A'
                is_part = prefix == 'P'
                is_comm = not (is_ass or is_part)
            # Determine parent number
            parts_num = num.split('.')
            parent_num = '.'.join(parts_num[:-1]) if len(parts_num) > 1 else None
            parent_struct = struct_map.get(parent_num)
            # Reset existing_global for this iteration.  This variable will be set
            # when a new structure is created based on a canonical existing node.
            existing_global = None
            # Create structure: allow multiple structures with the same name.  If a
            # non‑draft structure with the same name and typology exists, we will
            # copy its attributes onto the new structure after creation.
            struct = struct_map.get(num)
            if not struct:
                struct = Structure(name=code, type_id=stype.id,
                                   parent_id=parent_struct.id if parent_struct else None,
                                   flag_assembly=is_ass,
                                   flag_part=is_part,
                                   flag_commercial=is_comm)
                db.session.add(struct)
                db.session.flush()
                # Copy details from an existing global structure with the same name and typology
                existing_global = (Structure.query
                                   .filter_by(name=code,
                                              flag_assembly=is_ass,
                                              flag_part=is_part,
                                              flag_commercial=is_comm)
                                   .filter_by(table_assieme=False,
                                              table_parte=False,
                                              table_commerciale=False)
                                   .filter(Structure.id != struct.id)
                                   .first())
                if existing_global:
                    # Ensure the canonical existing structure is linked to a
                    # ComponentMaster so that attachments and attributes are
                    # stored centrally.  This call is safe to invoke multiple
                    # times and will create or reuse a master for the given
                    # code.
                    try:
                        ensure_component_master_for_structure(existing_global)
                    except Exception:
                        pass
                    attrs = [
                        'weight', 'processing_type', 'work_phase_id', 'supplier_id', 'work_center_id',
                        'standard_time', 'lead_time_theoretical', 'lead_time_real', 'description', 'notes',
                        'price_per_unit', 'minimum_order_qty', 'processing_cost'
                    ]
                    for attr in attrs:
                        val = getattr(existing_global, attr)
                        if val not in (None, ''):
                            setattr(struct, attr, val)
                    # persist updates
                    try:
                        db.session.flush()
                    except Exception:
                        db.session.rollback()
                    # After copying attributes from the canonical structure, replicate its
                    # attachments (image and document directories) to this newly created
                    # structure.  Without this, images and documents uploaded for one
                    # node would not appear on other nodes with the same code.
                    try:
                        # Replicate image from existing_global to new struct
                        upload_dir = os.path.join(current_app.static_folder, 'uploads')
                        if os.path.isdir(upload_dir):
                            prefix_src = f"sn_{existing_global.id}_"
                            src_file = None
                            for _f in os.listdir(upload_dir):
                                if _f.startswith(prefix_src):
                                    src_file = _f
                                    break
                            if src_file:
                                suffix = src_file[len(prefix_src):]
                                src_path = os.path.join(upload_dir, src_file)
                                dest_name = f"sn_{struct.id}_{suffix}"
                                dest_path = os.path.join(upload_dir, dest_name)
                                try:
                                    shutil.copyfile(src_path, dest_path)
                                except Exception:
                                    pass
                        # Replicate documents directory from existing_global to struct
                        def _safe_name(n: str) -> str:
                            return secure_filename(n) or 'unnamed'
                        type_dir_src = _safe_name(existing_global.type.name)
                        parts_src: list[str] = []
                        def _collect_src(node_obj):
                            if node_obj.parent:
                                _collect_src(node_obj.parent)
                            parts_src.append(_safe_name(node_obj.name))
                        _collect_src(existing_global)
                        base_src = os.path.join(current_app.static_folder, 'tmp_structures', type_dir_src, *parts_src)
                        if os.path.isdir(base_src):
                            type_dir_dst = _safe_name(struct.type.name)
                            parts_dst: list[str] = []
                            def _collect_dst(node_obj):
                                if node_obj.parent:
                                    _collect_dst(node_obj.parent)
                                parts_dst.append(_safe_name(node_obj.name))
                            _collect_dst(struct)
                            base_dst = os.path.join(current_app.static_folder, 'tmp_structures', type_dir_dst, *parts_dst)
                            try:
                                if os.path.isdir(base_dst):
                                    shutil.rmtree(base_dst)
                                shutil.copytree(base_src, base_dst)
                            except Exception:
                                pass
                    except Exception:
                        pass
                struct_map[num] = struct
                # -----------------------------------------------------------------
                # After creating a new structure (and copying attributes/attachments
                # from an existing global structure when applicable), ensure the
                # structure is linked to a ComponentMaster.  This call assigns
                # struct.component_id and copies any attachments from the
                # structure into the master directories.  If a master already
                # exists for this code (e.g. P001-025-001B) it will be reused.
                try:
                    ensure_component_master_for_structure(struct)
                except Exception:
                    # Do not block import on master assignment errors
                    pass
                # -----------------------------------------------------------------
                # If no existing global structure was used (meaning this is the
                # first occurrence of this code), assign weight and other
                # basic attributes from the BOM to the structure.  This makes
                # subsequent imports consistent (they will copy values from
                # the canonical structure).  Only assign when the value is
                # provided; do not overwrite existing values on repeated rows.
                if existing_global is None:
                    # Assign weight for parts and commercial items when available
                    if (is_part or is_comm) and weight_kg is not None:
                        # Only set if not already defined
                        if getattr(struct, 'weight', None) in (None, 0):
                            try:
                                struct.weight = weight_kg
                            except Exception:
                                pass
                    # Assign description for parts and assemblies
                    if desc:
                        try:
                            # Assemblies store description in notes on component but we
                            # can store it as description on structure to aid
                            # consistency across imports
                            struct.description = desc
                        except Exception:
                            pass
            # Create or update product component
            comp = ProductComponent.query.filter_by(product_id=product.id, structure_id=struct.id).first()
            if not comp:
                comp = ProductComponent(product_id=product.id, structure_id=struct.id, quantity=qty)
                db.session.add(comp)
            else:
                comp.quantity = qty
            # Always assign the component master id from the structure (if any).
            try:
                if struct.component_id:
                    comp.component_id = struct.component_id
            except Exception:
                pass
            # If this structure comes from an existing global node (meaning
            # ``existing_global`` was set and used), copy all attributes and
            # notes from the structure to the product component.  This ensures
            # that parts imported with codes already present inherit the
            # canonical definition (weight, processing, cycles, etc.).  If
            # ``existing_global`` is None, fall back to BOM values for the
            # component fields.
            if existing_global:
                # Use structure attributes for all categories
                comp.weight = struct.weight
                comp.processing_type = struct.processing_type
                comp.work_phase_id = struct.work_phase_id
                comp.supplier_id = struct.supplier_id
                comp.work_center_id = struct.work_center_id
                comp.standard_time = struct.standard_time
                comp.lead_time_theoretical = struct.lead_time_theoretical
                comp.lead_time_real = struct.lead_time_real
                comp.price_per_unit = struct.price_per_unit
                comp.minimum_order_qty = struct.minimum_order_qty
                comp.processing_cost = struct.processing_cost
                comp.description = struct.description
                comp.notes = struct.notes
            else:
                # Populate component fields based on BOM and category
                if is_part:
                    comp.weight = weight_kg
                    comp.processing_cost = cost
                    # For parts, store BOM description in description and keep notes blank
                    comp.description = desc or None
                    comp.notes = None
                elif is_comm:
                    comp.weight = weight_kg
                    comp.price_per_unit = cost
                    if material:
                        comp.description = f"{desc} {material}".strip() or None
                    else:
                        comp.description = desc or None
                    comp.notes = None
                else:
                    # Assemblies: store BOM description in description instead of notes.  When
                    # importing assemblies, the BOM often includes a description which
                    # should be visible in the warehouse views.  Assign it to the
                    # description field and leave notes empty.
                    comp.description = desc or None
                    comp.notes = None
            # Add material to dictionary
            if material:
                mc = MaterialCost.query.filter_by(material=material).first()
                if not mc:
                    mc = MaterialCost(material=material, cost_eur=None)
                    db.session.add(mc)
        db.session.commit()
        flash('Distinta importata con successo.', 'success')
        return redirect(url_for('products.detail', id=product.id))
    return render_template('admin/import_structure.html')

# -------------------------------------------------------------------------
# Import images by archive
@admin_bp.route('/import-images', methods=['GET', 'POST'])
@login_required
def import_images():
    """Importa un archivio di immagini e associa ogni file ad un componente.

    L'utente carica un file ZIP contenente immagini.  Ogni immagine deve
    essere nominata con il codice del componente (campo "Num. parte").  Se
    esiste un ComponentMaster corrispondente, l'immagine verrà salvata
    utilizzando il prefisso ``cm_<master.id>_``; in caso contrario verrà
    utilizzato il prefisso ``sn_<structure.id>_``.  Le immagini non
    riconosciute (senza struttura corrispondente) verranno ignorate.
    """
    admin_required()
    if request.method == 'POST':
        file = request.files.get('image_archive')
        if not file or not file.filename:
            flash('Seleziona un archivio di immagini.', 'warning')
            return redirect(url_for('admin.import_images'))
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext != '.zip':
            flash('Formato archivio non supportato. È richiesto un file ZIP.', 'danger')
            return redirect(url_for('admin.import_images'))
        try:
            with tempfile.TemporaryDirectory() as tmpdir:
                archive_path = os.path.join(tmpdir, filename)
                file.save(archive_path)
                # Extract contents of the archive
                try:
                    with zipfile.ZipFile(archive_path, 'r') as zf:
                        zf.extractall(tmpdir)
                except Exception:
                    flash("Impossibile estrarre l'archivio.", 'danger')
                    return redirect(url_for('admin.import_images'))
                # Helper to validate image extensions
                def _allowed_image(fname: str) -> bool:
                    return '.' in fname and fname.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'bmp'}
                upload_dir = os.path.join(current_app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                imported_count = 0
                unmatched: list[str] = []
                # Walk through extracted files
                for root, dirs, files in os.walk(tmpdir):
                    for fname in files:
                        # Skip the original archive file itself
                        if root == tmpdir and fname == filename:
                            continue
                        if not _allowed_image(fname):
                            continue
                        code = os.path.splitext(fname)[0].strip()
                        if not code:
                            continue
                        # Find structures by name (code)
                        structs = Structure.query.filter_by(name=code).all()
                        if not structs:
                            unmatched.append(fname)
                            continue
                        src_path = os.path.join(root, fname)
                        for s in structs:
                            # Ensure a master exists for this structure
                            master_obj = ensure_component_master_for_structure(s)
                            if master_obj:
                                dest_prefix = f"cm_{master_obj.id}_"
                            else:
                                dest_prefix = f"sn_{s.id}_"
                            dest_name = dest_prefix + secure_filename(fname)
                            dest_path = os.path.join(upload_dir, dest_name)
                            try:
                                shutil.copyfile(src_path, dest_path)
                                imported_count += 1
                            except Exception:
                                pass
                # Commit any masters created during import
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                # Feedback messages
                if imported_count > 0:
                    flash(f'Importate {imported_count} immagini.', 'success')
                if unmatched:
                    preview = ', '.join(unmatched[:5])
                    if len(unmatched) > 5:
                        preview += ', ...'
                    flash(f'Alcune immagini non sono state associate: {preview}', 'warning')
                return redirect(url_for('admin.structures'))
        except Exception:
            flash("Errore durante l'importazione delle immagini.", 'danger')
            return redirect(url_for('admin.import_images'))
    # GET request: render the upload form
    return render_template('admin/import_images.html')

# ---------------------------------------------------------
# Routes for editing structure types and nodes
#
# Administrators can modify existing structure types and structure nodes.
# Each edit endpoint presents a form pre‑populated with the current values.
@admin_bp.route('/structures/type/<int:type_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_structure_type(type_id: int):
    """Edit an existing StructureType.

    This view allows administrators to change the name, description and typology
    flags (assembly, part, commercial) of a structure type.  On successful
    submission the changes are persisted and the user is redirected back to
    the structures overview."""
    admin_required()
    stype = StructureType.query.get_or_404(type_id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        # Read typology from radio input.  When the typology field is absent (e.g. the
        # form does not include the radio buttons), set to None so that the
        # existing flags remain unchanged.  Otherwise expect one of
        # 'assembly', 'part', or 'commercial'.
        typology = request.form.get('typology') or None
        is_ass = typology == 'assembly'
        is_part = typology == 'part'
        is_comm = typology == 'commercial'
        if not name:
            flash('Il nome è obbligatorio.', 'danger')
        else:
            # Ensure name uniqueness among other types
            existing = (
                StructureType.query
                .filter(StructureType.name == name, StructureType.id != stype.id)
                .first()
            )
            if existing:
                flash('Esiste già un tipo con questo nome.', 'warning')
            else:
                stype.name = name
                stype.description = description
                # Only update typology flags when the form explicitly provided a selection
                if typology:
                    stype.is_assembly = is_ass
                    stype.is_part = is_part
                    stype.is_commercial = is_comm
                # Update default attributes
                def_phase_id = request.form.get('default_work_phase_id') or None
                def_proc_type = request.form.get('default_processing_type') or None
                def_supplier_id = request.form.get('default_supplier_id') or None
                def_center_id = request.form.get('default_work_center_id') or None
                def_weight = request.form.get('default_weight') or None
                def_std_hours = request.form.get('default_standard_time') or None
                def_lead_theo_days = request.form.get('default_lead_time_theoretical') or None
                def_lead_real_days = request.form.get('default_lead_time_real') or None
                def_desc = request.form.get('default_description') or None
                def_notes = request.form.get('default_notes') or None
                def_price = request.form.get('default_price_per_unit') or None
                def_min_qty = request.form.get('default_minimum_order_qty') or None
                def_proc_cost = request.form.get('default_processing_cost') or None
                stype.default_work_phase_id = int(def_phase_id) if def_phase_id else None
                stype.default_processing_type = def_proc_type
                stype.default_supplier_id = int(def_supplier_id) if def_supplier_id else None
                stype.default_work_center_id = int(def_center_id) if def_center_id else None
                try:
                    stype.default_weight = float(def_weight) if def_weight else None
                except (ValueError, TypeError):
                    stype.default_weight = None
                # Standard time in minutes
                std_minutes = None
                try:
                    std_hours = float(def_std_hours) if def_std_hours else None
                    std_minutes = std_hours * 60.0 if std_hours is not None else None
                except (ValueError, TypeError):
                    std_minutes = None
                stype.default_standard_time = std_minutes
                # Lead times to minutes
                def to_minutes(days_str):
                    try:
                        return float(days_str) * 1440.0
                    except Exception:
                        return None
                stype.default_lead_time_theoretical = to_minutes(def_lead_theo_days)
                stype.default_lead_time_real = to_minutes(def_lead_real_days)
                stype.default_description = def_desc
                stype.default_notes = def_notes
                try:
                    stype.default_price_per_unit = float(def_price) if def_price else None
                except (ValueError, TypeError):
                    stype.default_price_per_unit = None
                try:
                    stype.default_minimum_order_qty = int(def_min_qty) if def_min_qty else None
                except (ValueError, TypeError):
                    stype.default_minimum_order_qty = None
                # Determine default processing cost: prefer user-provided value
                stype.default_processing_cost = None
                if def_proc_cost:
                    try:
                        stype.default_processing_cost = float(def_proc_cost)
                    except (ValueError, TypeError):
                        stype.default_processing_cost = None
                if stype.default_processing_cost is None:
                    if def_proc_type == 'internal' and std_minutes is not None and stype.default_work_center_id:
                        wc = WorkCenter.query.get(stype.default_work_center_id)
                        if wc and wc.hourly_cost is not None:
                            stype.default_processing_cost = (std_minutes / 60.0) * wc.hourly_cost
                db.session.commit()
                # Propagate updated default attributes to structures of this type
                structures = stype.nodes.all()
                for s in structures:
                    # Only update attributes that are not explicitly set on the node
                    if s.weight is None and stype.default_weight is not None:
                        s.weight = stype.default_weight
                    if (not s.processing_type) and stype.default_processing_type:
                        s.processing_type = stype.default_processing_type
                    if s.work_phase_id is None and stype.default_work_phase_id:
                        s.work_phase_id = stype.default_work_phase_id
                    if s.supplier_id is None and stype.default_supplier_id:
                        s.supplier_id = stype.default_supplier_id
                    if s.work_center_id is None and stype.default_work_center_id:
                        s.work_center_id = stype.default_work_center_id
                    if s.standard_time is None and stype.default_standard_time is not None:
                        s.standard_time = stype.default_standard_time
                    if s.lead_time_theoretical is None and stype.default_lead_time_theoretical is not None:
                        s.lead_time_theoretical = stype.default_lead_time_theoretical
                    if s.lead_time_real is None and stype.default_lead_time_real is not None:
                        s.lead_time_real = stype.default_lead_time_real
                    if not s.description and stype.default_description:
                        s.description = stype.default_description
                    if not s.notes and stype.default_notes:
                        s.notes = stype.default_notes
                    if s.price_per_unit is None and stype.default_price_per_unit is not None:
                        s.price_per_unit = stype.default_price_per_unit
                    if s.minimum_order_qty is None and stype.default_minimum_order_qty is not None:
                        s.minimum_order_qty = stype.default_minimum_order_qty
                    # Compute processing cost for internal processes
                    s.processing_cost = None
                    if s.processing_type == 'internal' and s.standard_time is not None and s.work_center_id:
                        wc_i = WorkCenter.query.get(s.work_center_id)
                        if wc_i and wc_i.hourly_cost is not None:
                            s.processing_cost = (s.standard_time / 60.0) * wc_i.hourly_cost
                    elif s.processing_type == 'external' and stype.default_processing_cost is not None:
                        s.processing_cost = stype.default_processing_cost
                    # Propagate to product components linked to this structure
                    comps = ProductComponent.query.filter_by(structure_id=s.id).all()
                    for comp in comps:
                        if comp.weight is None:
                            comp.weight = s.weight
                        if not comp.processing_type:
                            comp.processing_type = s.processing_type
                        if comp.work_phase_id is None:
                            comp.work_phase_id = s.work_phase_id
                        if comp.supplier_id is None:
                            comp.supplier_id = s.supplier_id
                        if comp.work_center_id is None:
                            comp.work_center_id = s.work_center_id
                        if comp.standard_time is None:
                            comp.standard_time = s.standard_time
                        if comp.lead_time_theoretical is None:
                            comp.lead_time_theoretical = s.lead_time_theoretical
                        if comp.lead_time_real is None:
                            comp.lead_time_real = s.lead_time_real
                        if not comp.description:
                            comp.description = s.description
                        if not comp.notes:
                            comp.notes = s.notes
                        if comp.price_per_unit is None:
                            comp.price_per_unit = s.price_per_unit
                        if comp.minimum_order_qty is None:
                            comp.minimum_order_qty = s.minimum_order_qty
                        # Update processing cost on component
                        comp.processing_cost = s.processing_cost
                db.session.commit()
                flash('Tipo aggiornato.', 'success')
                return redirect(url_for('admin.structures'))
    # Provide dictionary lists for definisci selects
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()
    return render_template('admin/edit_structure_type.html',
                           type=stype,
                           suppliers=suppliers,
                           work_centers=work_centers,
                           work_phases=work_phases)


@admin_bp.route('/structures/node/<int:node_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_structure_node(node_id: int):
    """Edit an existing Structure node.

    This view lets administrators update the node name, assign it to a type,
    select a parent node, and adjust typology flags.  On submission the
    changes are saved and the admin is redirected to the structures page."""
    admin_required()
    node = Structure.query.get_or_404(node_id)
    types = StructureType.query.order_by(StructureType.name.asc()).all()
    nodes_by_type = {t.id: t.nodes.order_by(Structure.name.asc()).all() for t in types}

    if request.method == 'POST':
        name = request.form.get('node_name', '').strip()
        type_id_val = request.form.get('type_id')
        parent_id_val = request.form.get('parent_id') or None
        # Read typology from radio selection
        node_typology = request.form.get('node_typology') or ''
        flag_ass = node_typology == 'assembly'
        flag_part = node_typology == 'part'
        flag_comm = node_typology == 'commercial'
        if not name or not type_id_val:
            flash('Compila tutti i campi obbligatori.', 'danger')
        else:
            if not node_typology:
                flash('Seleziona una tipologia per il nodo (Assieme, Parte o Parte a commercio).', 'danger')
            else:
                node.name = name
                node.type_id = int(type_id_val)
                if parent_id_val and int(parent_id_val) != node.id:
                    node.parent_id = int(parent_id_val)
                else:
                    node.parent_id = None
                node.flag_assembly = flag_ass
                node.flag_part = flag_part
                node.flag_commercial = flag_comm
                # Update default attributes from definisci section
                node_phase_id = request.form.get('node_work_phase_id') or None
                node_proc_type = request.form.get('node_processing_type') or None
                node_supplier_id = request.form.get('node_supplier_id') or None
                node_center_id = request.form.get('node_work_center_id') or None
                node_weight = request.form.get('node_weight') or None
                node_std_hours = request.form.get('node_standard_time') or None
                node_lead_theo = request.form.get('node_lead_time_theoretical') or None
                node_lead_real = request.form.get('node_lead_time_real') or None
                node_desc = request.form.get('node_description') or None
                node_notes = request.form.get('node_notes') or None
                node_price = request.form.get('node_price_per_unit') or None
                node_min_qty = request.form.get('node_minimum_order_qty') or None
                node_proc_cost = request.form.get('node_processing_cost') or None
                node.work_phase_id = int(node_phase_id) if node_phase_id else None
                node.processing_type = node_proc_type
                node.supplier_id = int(node_supplier_id) if node_supplier_id else None
                node.work_center_id = int(node_center_id) if node_center_id else None
                try:
                    node.weight = float(node_weight) if node_weight else None
                except (ValueError, TypeError):
                    node.weight = None
                # Standard time to minutes
                std_minutes = None
                try:
                    std_h = float(node_std_hours) if node_std_hours else None
                    std_minutes = std_h * 60.0 if std_h is not None else None
                except (ValueError, TypeError):
                    std_minutes = None
                node.standard_time = std_minutes
                # Lead times to minutes
                def to_minutes(days_str):
                    try:
                        return float(days_str) * 1440.0
                    except Exception:
                        return None
                node.lead_time_theoretical = to_minutes(node_lead_theo)
                node.lead_time_real = to_minutes(node_lead_real)
                node.description = node_desc
                node.notes = node_notes
                try:
                    node.price_per_unit = float(node_price) if node_price else None
                except (ValueError, TypeError):
                    node.price_per_unit = None
                try:
                    node.minimum_order_qty = int(node_min_qty) if node_min_qty else None
                except (ValueError, TypeError):
                    node.minimum_order_qty = None
                # Determine processing cost: prefer provided value, else compute internal
                node.processing_cost = None
                if node_proc_cost:
                    try:
                        node.processing_cost = float(node_proc_cost)
                    except (ValueError, TypeError):
                        node.processing_cost = None
                if node.processing_cost is None:
                    if node_proc_type == 'internal' and std_minutes is not None and node.work_center_id:
                        wc = WorkCenter.query.get(node.work_center_id)
                        if wc and wc.hourly_cost is not None:
                            node.processing_cost = (std_minutes / 60.0) * wc.hourly_cost
                    # External: no automatic cost
                db.session.commit()
                # Propagate updated default attributes to existing product components associated with this node
                comps = ProductComponent.query.filter_by(structure_id=node.id).all()
                for comp in comps:
                    comp.weight = node.weight
                    comp.processing_type = node.processing_type
                    comp.work_phase_id = node.work_phase_id
                    comp.supplier_id = node.supplier_id
                    comp.work_center_id = node.work_center_id
                    comp.standard_time = node.standard_time
                    comp.lead_time_theoretical = node.lead_time_theoretical
                    comp.lead_time_real = node.lead_time_real
                    comp.description = node.description
                    comp.notes = node.notes
                    comp.price_per_unit = node.price_per_unit
                    comp.minimum_order_qty = node.minimum_order_qty
                    comp.processing_cost = node.processing_cost
                db.session.commit()
                # Propagate updated attributes to any other structure nodes with the same name and typology.
                # This ensures that editing one copy of a node updates all duplicates across types.
                duplicates = (Structure.query
                              .filter_by(name=node.name,
                                         flag_assembly=node.flag_assembly,
                                         flag_part=node.flag_part,
                                         flag_commercial=node.flag_commercial)
                              .filter(Structure.id != node.id)
                              .all())
                if duplicates:
                    attrs = [
                        'type_id', 'parent_id',
                        'weight', 'processing_type', 'work_phase_id', 'supplier_id', 'work_center_id',
                        'standard_time', 'lead_time_theoretical', 'lead_time_real', 'description', 'notes',
                        'price_per_unit', 'minimum_order_qty', 'processing_cost'
                    ]
                    for dup in duplicates:
                        for attr in attrs:
                            setattr(dup, attr, getattr(node, attr))
                        # Propagate attributes to product components of the duplicate structure
                        dup_comps = ProductComponent.query.filter_by(structure_id=dup.id).all()
                        for comp in dup_comps:
                            comp.weight = node.weight
                            comp.processing_type = node.processing_type
                            comp.work_phase_id = node.work_phase_id
                            comp.supplier_id = node.supplier_id
                            comp.work_center_id = node.work_center_id
                            comp.standard_time = node.standard_time
                            comp.lead_time_theoretical = node.lead_time_theoretical
                            comp.lead_time_real = node.lead_time_real
                            comp.description = node.description
                            comp.notes = node.notes
                            comp.price_per_unit = node.price_per_unit
                            comp.minimum_order_qty = node.minimum_order_qty
                            comp.processing_cost = node.processing_cost
                    db.session.commit()
                # -----------------------------------------------------------------
                # Ensure a master component exists for the updated node and its duplicates.
                # This will assign component_id and copy attachments.  Execute in a
                # try block to avoid breaking the flow if anything fails.
                try:
                    ensure_component_master_for_structure(node)
                    if duplicates:
                        for dup_node in duplicates:
                            ensure_component_master_for_structure(dup_node)
                except Exception:
                    pass
                flash('Nodo aggiornato.', 'success')
                return redirect(url_for('admin.structures'))
    # Provide dictionary lists for definisci selects
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    work_centers = WorkCenter.query.order_by(WorkCenter.name.asc()).all()
    work_phases = WorkPhase.query.order_by(WorkPhase.name.asc()).all()
    return render_template('admin/edit_structure_node.html',
                           node=node,
                           types=types,
                           nodes_by_type=nodes_by_type,
                           suppliers=suppliers,
                           work_centers=work_centers,
                           work_phases=work_phases)


# -----------------------------------------------------------------------------
# Eliminazione dei tipi di struttura
#
# Gli amministratori possono eliminare un tipo di struttura principale.  Prima
# dell'eliminazione viene mostrata una pagina di conferma per evitare rimozioni
# accidentali.  Quando un tipo viene eliminato, anche tutti i nodi associati
# vengono rimossi, unitamente ai componenti di prodotto che li referenziano.

@admin_bp.route('/structures/type/<int:type_id>/delete', methods=['GET', 'POST'])
@login_required
def delete_structure_type(type_id: int):
    """Elimina un tipo di struttura e tutti i suoi nodi.

    Questo endpoint supporta sia GET che POST: con GET viene mostrata una
    pagina di conferma, con POST avviene l'eliminazione effettiva.  Se
    l'eliminazione ha successo si torna alla pagina delle strutture con un
    messaggio di conferma.
    """
    admin_required()
    stype = StructureType.query.get_or_404(type_id)
    if request.method == 'POST':
        # Prima rimuove tutti i nodi associati al tipo, insieme ai relativi
        # componenti di prodotto.  Poiché le relazioni non sono configurate
        # con cascade delete sul database, effettuiamo l'eliminazione manualmente.
        nodes = stype.nodes.all()
        for node in nodes:
            # Elimina eventuali componenti che referenziano il nodo
            comps = ProductComponent.query.filter_by(structure_id=node.id).all()
            for comp in comps:
                db.session.delete(comp)
            # Elimina il nodo stesso.  I figli verranno eliminati dal ciclo
            # successivo perché ``nodes`` contiene tutti i nodi del tipo.
            db.session.delete(node)
        # Elimina eventuali campi personalizzati associati al tipo (TypeField)
        for field in stype.fields:
            db.session.delete(field)
        # Infine elimina il tipo
        db.session.delete(stype)
        db.session.commit()
        flash('Tipo eliminato con successo.', 'success')
        return redirect(url_for('admin.structures'))
    # GET: visualizza pagina di conferma
    return render_template('admin/delete_structure_type.html', type=stype)

@admin_bp.route('/structures/node/<int:node_id>/delete', methods=['GET', 'POST'])
@login_required
def delete_structure_node(node_id: int):
    """Elimina un singolo nodo di struttura e tutti i suoi figli.

    Simile alla cancellazione del tipo, questo endpoint mostra una pagina di
    conferma con GET e procede all'eliminazione con POST.  Eliminare un
    nodo comporta anche la rimozione di tutti i componenti di prodotto
    associati e di tutti i nodi figli nella gerarchia.  La rimozione è
    irreversibile.
    """
    admin_required()
    node = Structure.query.get_or_404(node_id)
    if request.method == 'POST':
        # Delete recursively children and associated product components
        def _delete_recursively(n: Structure):
            # First delete descendants
            for child in n.children:
                _delete_recursively(child)
            # Delete product components referencing this structure
            comps = ProductComponent.query.filter_by(structure_id=n.id).all()
            for comp in comps:
                db.session.delete(comp)
            # Delete the node itself
            db.session.delete(n)
        try:
            _delete_recursively(node)
            db.session.commit()
            flash('Nodo eliminato con successo.', 'success')
        except Exception:
            db.session.rollback()
            flash("Errore durante l'eliminazione del nodo.", 'danger')
        return redirect(url_for('admin.structures'))
    # GET: render confirmation template
    return render_template('admin/delete_structure_node.html', node=node)


